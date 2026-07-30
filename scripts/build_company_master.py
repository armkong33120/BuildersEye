#!/usr/bin/env python3
"""Build company-master.xlsx with 6 business sheets (LOOP 24)."""
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

APP_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = APP_ROOT / "src" / "data"
OUTPUT_DIR = DATA_DIR / "hr_onedrive_demo"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))

def write_header(ws, headers, row=1):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = WRAP_ALIGN; c.border = THIN_BORDER

def safe_val(v):
    if isinstance(v, (list, dict)):
        return json.dumps(v, ensure_ascii=False)
    return v

def write_row(ws, rn, vals):
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=rn, column=ci, value=safe_val(v))
        c.font = BODY_FONT; c.alignment = WRAP_ALIGN; c.border = THIN_BORDER

def auto_width(ws, min_w=8, max_w=50):
    for col in ws.columns:
        ml = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(min_w, min(max_w, ml + 2))

sheets = [
    ("Product_Catalog", "product-catalog.json"),
    ("Revenue_By_Product", "revenue-by-product.json"),
    ("Customer_Portfolio", "customer-portfolio.json"),
    ("Department_PnL", "department-pnl.json"),
    ("Sales_Pipeline", "sales-pipeline.json"),
    ("Operating_Expenses", "company-operating-expenses.json"),
]

wb = openpyxl.Workbook()
first = True
for sheet_name, fname in sheets:
    path = DATA_DIR / fname
    if not path.exists():
        print(f"  WARN: {fname} not found, skipping {sheet_name}")
        continue
    data = json.loads(path.read_text("utf-8"))
    if first:
        ws = wb.active
        ws.title = sheet_name
        first = False
    else:
        ws = wb.create_sheet(sheet_name)
    if data:
        keys = list(data[0].keys())
        write_header(ws, keys)
        for i, row in enumerate(data, 2):
            write_row(ws, i, [row.get(k, "") for k in keys])
        auto_width(ws)
    print(f"  Sheet '{sheet_name}': {len(data)} rows")

fp = OUTPUT_DIR / "company-master.xlsx"
wb.save(fp)
print(f"\nSaved: {fp}")