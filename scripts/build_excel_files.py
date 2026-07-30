#!/usr/bin/env python3
"""Generate 150 employee Excel files (23 sheets) + 1 Master Index + Validation Report."""
import json, os, random, zipfile
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

random.seed(20260707)
APP_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = APP_ROOT / "src" / "data"
OUTPUT_DIR = DATA_DIR / "hr_onedrive_demo"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

identity = json.loads((DATA_DIR / "identity-graph.json").read_text("utf-8"))
careers = json.loads((DATA_DIR / "career-story-plan.json").read_text("utf-8"))
kpi_all_input = json.loads((DATA_DIR / "kpi-okr-history.json").read_text("utf-8"))
assignments_all = json.loads((DATA_DIR / "project-assignments.json").read_text("utf-8"))
collabs_all = json.loads((DATA_DIR / "collaboration-graph.json").read_text("utf-8"))
warnings_all = json.loads((DATA_DIR / "warning-history.json").read_text("utf-8"))
learning_all = json.loads((DATA_DIR / "learning-history.json").read_text("utf-8"))
master_index = json.loads((DATA_DIR / "master-index.json").read_text("utf-8"))
# LOOP 20 datasets
it_assets_all = json.loads((DATA_DIR / "it-assets.json").read_text("utf-8"))
it_tickets_all = json.loads((DATA_DIR / "it-tickets.json").read_text("utf-8"))
sw_licenses_all = json.loads((DATA_DIR / "software-licenses.json").read_text("utf-8"))
salary_all = json.loads((DATA_DIR / "salary-history.json").read_text("utf-8"))
attendance_all = json.loads((DATA_DIR / "attendance-record.json").read_text("utf-8"))
# LOOP 22: 11 new HCM datasets
feedback_all = json.loads((DATA_DIR / "feedback-360.json").read_text("utf-8"))
skill_all = json.loads((DATA_DIR / "skill-matrix.json").read_text("utf-8"))
succession_all = json.loads((DATA_DIR / "succession-planning.json").read_text("utf-8"))
benefits_all = json.loads((DATA_DIR / "benefit-claims.json").read_text("utf-8"))
expenses_all = json.loads((DATA_DIR / "expense-reports.json").read_text("utf-8"))
grievances_all = json.loads((DATA_DIR / "grievance-log.json").read_text("utf-8"))
compliance_all = json.loads((DATA_DIR / "compliance-mandates.json").read_text("utf-8"))
onboarding_all = json.loads((DATA_DIR / "onboarding-journey.json").read_text("utf-8"))
engagement_all = json.loads((DATA_DIR / "employee-engagement.json").read_text("utf-8"))
security_all = json.loads((DATA_DIR / "physical-security.json").read_text("utf-8"))
timesheet_all = json.loads((DATA_DIR / "timesheet-log.json").read_text("utf-8"))

employees_by_pk = {e["pk"]: e for e in identity["identities"]}
car_by_pk = {c["employeeId"]: c for c in careers}
warn_raw_by_pk = defaultdict(list)
for w in warnings_all:
    warn_raw_by_pk[w["employeeId"]].append(w)

# ---------- KPI POST-PROCESSING: inject A-band and E-band ----------
a_candidates = {1, 2, 3}
for pk in range(4, 151):
    emp = employees_by_pk[pk]
    car = car_by_pk.get(pk, {})
    warns = warn_raw_by_pk.get(pk, [])
    high_warns = sum(1 for w in warns if w["severity"] in ("High", "Critical"))
    if car.get("tenureMonths", 0) > 60 and high_warns == 0 and emp["hierarchyDepth"] <= 2:
        a_candidates.add(pk)
    if len(a_candidates) >= 6:
        break
a_band_pks = set(random.sample(sorted(a_candidates), min(6, len(a_candidates))))

e_candidates = set()
for pk in range(1, 151):
    warns = warn_raw_by_pk.get(pk, [])
    formal_warns = sum(1 for w in warns if w.get("formalWarning"))
    emp = employees_by_pk[pk]
    if formal_warns >= 1 and emp["hierarchyDepth"] >= 3:
        e_candidates.add(pk)
    if len(e_candidates) >= 8:
        break
e_band_pks = set(random.sample(sorted(e_candidates), min(5, len(e_candidates))) if e_candidates else set())

kpi_all = []
for k in kpi_all_input:
    pk = k["employeeId"]
    band = k["performanceBand"]
    if pk in a_band_pks and band in ("Exceeds (B)", "Meets (C)"):
        kpis_for_pk = sorted([x for x in kpi_all_input if x["employeeId"] == pk], key=lambda x: x["reviewPeriod"])
        if kpis_for_pk and k["reviewPeriod"] == kpis_for_pk[-1]["reviewPeriod"]:
            k = dict(k)
            k["kpiScore"] = round(random.uniform(4.5, 4.8), 1)
            k["okrScore"] = round(k["kpiScore"] - random.uniform(0.2, 0.5), 1)
            k["performanceBand"] = "Exceptional (A)"
            k["strongArea"] = "Strategic Leadership"
            k["weakArea"] = "Work-Life Balance"
            k["managerFeedback"] = "ผลงานโดดเด่นเกินความคาดหมายในทุกด้าน — เป็นแบบอย่างให้องค์กร"
    if pk in e_band_pks and band in ("Meets (C)", "Below (D)"):
        kpis_for_pk = sorted([x for x in kpi_all_input if x["employeeId"] == pk], key=lambda x: x["reviewPeriod"])
        if len(kpis_for_pk) >= 3 and k["reviewPeriod"] == kpis_for_pk[-2]["reviewPeriod"]:
            k = dict(k)
            k["kpiScore"] = round(random.uniform(1.0, 1.4), 1)
            k["okrScore"] = round(k["kpiScore"] - random.uniform(0.1, 0.3), 1)
            k["performanceBand"] = "Unsatisfactory (E)"
            k["weakArea"] = "Performance Consistency"
            k["managerFeedback"] = "ต้องปรับปรุงอย่างเร่งด่วน — มีการกำหนด PIP และติดตามอย่างใกล้ชิด"
            k["improvementPlan"] = "กำหนด Performance Improvement Plan 60 วัน"
            k["followUpStatus"] = "In Progress"
    kpi_all.append(k)

kpi_by_pk = defaultdict(list)
for k in kpi_all:
    kpi_by_pk[k["employeeId"]].append(k)

assign_by_pk = defaultdict(list)
for a in assignments_all:
    assign_by_pk[a["employeeId"]].append(a)

collab_by_pk = defaultdict(list)
for c in collabs_all:
    collab_by_pk[c["employeeId"]].append(c)

warn_by_pk = defaultdict(list)
for w in warnings_all:
    warn_by_pk[w["employeeId"]].append(w)

learn_by_pk = defaultdict(list)
for l in learning_all:
    learn_by_pk[l["employeeId"]].append(l)

# LOOP 20: Group advanced data by employeeId
it_assets_by_pk = defaultdict(list)
for a in it_assets_all: it_assets_by_pk[a["employeeId"]].append(a)
it_tickets_by_pk = defaultdict(list)
for t in it_tickets_all: it_tickets_by_pk[t["employeeId"]].append(t)
sw_licenses_by_pk = defaultdict(list)
for l in sw_licenses_all: sw_licenses_by_pk[l["employeeId"]].append(l)
salary_by_pk = {}
for s in salary_all: salary_by_pk[s["employeeId"]] = s
attendance_by_pk = {}
for a in attendance_all: attendance_by_pk[a["employeeId"]] = a

# LOOP 22: Group 11 new datasets by employeeId
feedback_by_pk = {}
for f in feedback_all: feedback_by_pk[f["employeeId"]] = f
skill_by_pk = {}
for s in skill_all: skill_by_pk[s["employeeId"]] = s
succession_by_pk = {}
for s in succession_all: succession_by_pk[s["employeeId"]] = s
benefits_by_pk = {}
for b in benefits_all: benefits_by_pk[b["employeeId"]] = b
expenses_by_pk = {}
for e in expenses_all: expenses_by_pk[e["employeeId"]] = e
grievances_by_pk = defaultdict(list)
for g in grievances_all: grievances_by_pk[g["employeeId"]].append(g)
compliance_by_pk = defaultdict(list)
for c in compliance_all: compliance_by_pk[c["employeeId"]].append(c)
onboarding_by_pk = {}
for o in onboarding_all: onboarding_by_pk[o["employeeId"]] = o
engagement_by_pk = {}
for e in engagement_all: engagement_by_pk[e["employeeId"]] = e
security_by_pk = {}
for s in security_all: security_by_pk[s["employeeId"]] = s
timesheet_by_pk = {}
for t in timesheet_all: timesheet_by_pk[t["employeeId"]] = t

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
RED_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
GREEN_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

CONFIDENTIALITY_MAP = {"Executive": "Tier 1 - Strict", "HR & Admin": "Tier 1 - Strict",
                       "Finance & Accounting": "Tier 1 - Strict", "Legal": "Tier 1 - Strict", "IT": "Tier 2 - Sensitive"}

def write_header(ws, headers, row=1):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = WRAP_ALIGN; c.border = THIN_BORDER

def write_row(ws, rn, vals):
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=rn, column=ci, value=v)
        c.font = BODY_FONT; c.alignment = WRAP_ALIGN; c.border = THIN_BORDER

def auto_width(ws, min_w=8, max_w=40):
    for col in ws.columns:
        ml = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(min_w, min(max_w, ml + 2))

def add_title(ws, title, subtitle=""):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(row=1, column=1, value=title).font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
        ws.cell(row=2, column=1, value=subtitle).font = Font(name="Calibri", size=10, italic=True, color="666666")

def build_timeline(pk):
    emp = employees_by_pk[pk]; car = car_by_pk.get(pk, {})
    jd = car.get("joinDate", ""); pc = car.get("promotionCount", 0)
    sp = car.get("startingPosition", ""); ct = emp["jobTitle"]
    cs = car.get("currentPositionStartDate", ""); dept = emp["department"]
    events = []
    if jd: events.append({"date": jd, "eventType": "Hired", "title": sp, "department": dept, "notes": "Joined the company"})
    if pc > 0 and jd and cs:
        jdt = datetime.strptime(jd, "%Y-%m-%d"); cst = datetime.strptime(cs, "%Y-%m-%d")
        for i in range(pc):
            mid = jdt + (cst - jdt) * ((i + 1) / (pc + 1))
            events.append({"date": mid.strftime("%Y-%m-%d"), "eventType": "Promoted",
                "title": f"Intermediate Role {i+1}", "department": dept, "notes": f"Promotion #{i+1}"})
    if cs and cs != jd: events.append({"date": cs, "eventType": "Promoted", "title": ct, "department": dept, "notes": "Reached current position"})
    assigns = assign_by_pk.get(pk, [])
    for a in random.sample(assigns, min(3, len(assigns))) if assigns else []:
        events.append({"date": f"2024-Q{random.randint(1,4)}", "eventType": "ProjectAssigned", "title": a["projectId"], "department": dept, "notes": f"Role: {a['role']}"})
    if pc >= 2: events.append({"date": f"2023-Q{random.randint(1,4)}", "eventType": "CertificationEarned", "title": "Professional Certification", "department": dept, "notes": "Earned professional certification"})
    events.sort(key=lambda x: x["date"])
    return events

def gen_employee(pk):
    emp = employees_by_pk[pk]; code = emp["code"]; name = emp["name"]; dept = emp["department"]
    fp = OUTPUT_DIR / f"{code}_OneDrive_Profile.xlsx"
    wb = openpyxl.Workbook(); car = car_by_pk.get(pk, {})
    is_t1 = CONFIDENTIALITY_MAP.get(dept, "") in ("Tier 1 - Strict", "Tier 2 - Sensitive")
    sw = car.get("mainWeakness", "")
    if is_t1 and random.random() < 0.3: sw = "[Redacted - Tier 1 Confidentiality]"

    # Sheet 1: Employee_Profile
    ws1 = wb.active; ws1.title = "Employee_Profile"
    add_title(ws1, f"Employee Profile - {code} {name}", f"Department: {dept} | Generated: 2026-06-26")
    h1 = ["pk","code","name","department","jobTitle","roleGroup","managerCode","managerName",
          "email","mailAlias","licensePlan","mailboxQuotaGb","oneDriveQuotaGb","oneDriveUrl",
          "oneDriveOwner","mfaStatus","accountStatus","accountRisk","hierarchyDepth",
          "directReportCount","hireDate","tenureMonths","startingPosition",
          "currentPositionStartDate","promotionCount","mainStrength","mainWeakness",
          "learningTheme","retentionRisk","successionPotential"]
    write_header(ws1, h1, 3)
    write_row(ws1, 4, [pk,code,name,dept,emp["jobTitle"],emp.get("roleGroup",""),
        emp.get("managerCode",""),emp.get("managerName",""),emp["email"],
        emp.get("mailAlias",""),emp["licensePlan"],emp["mailboxQuotaGb"],
        emp["oneDriveQuotaGb"],emp["oneDriveUrl"],emp["oneDriveOwner"],
        emp["mfaStatus"],emp["accountStatus"],emp["accountRisk"],
        emp["hierarchyDepth"],emp.get("directReportCount",0),
        car.get("joinDate",""),car.get("tenureMonths",0),car.get("startingPosition",""),
        car.get("currentPositionStartDate",""),car.get("promotionCount",0),
        car.get("mainStrength",""),sw,car.get("learningTheme",""),
        car.get("retentionRisk",""),car.get("successionPotential","")])
    auto_width(ws1)

    # Sheet 2: Career_Timeline
    ws2 = wb.create_sheet("Career_Timeline"); add_title(ws2, f"Career Timeline - {name}")
    write_header(ws2, ["date","eventType","title","department","notes"], 3)
    for i,ev in enumerate(build_timeline(pk),4): write_row(ws2,i,[ev["date"],ev["eventType"],ev["title"],ev["department"],ev["notes"]])
    auto_width(ws2)

    # Sheet 3: KPI_OKR_History
    ws3 = wb.create_sheet("KPI_OKR_History"); add_title(ws3, f"KPI / OKR History - {name}")
    write_header(ws3, ["reviewPeriod","kpiScore","okrScore","performanceBand","strongArea","weakArea","managerFeedback","improvementPlan","followUpStatus"], 3)
    for i,k in enumerate(sorted(kpi_by_pk.get(pk,[]),key=lambda x:x["reviewPeriod"]),4):
        write_row(ws3,i,[k["reviewPeriod"],k["kpiScore"],k["okrScore"],k["performanceBand"],
            k["strongArea"],k["weakArea"],k["managerFeedback"],k.get("improvementPlan",""),k.get("followUpStatus","")])
        if k["performanceBand"] in ("Below (D)","Unsatisfactory (E)"):
            for ci in range(1,10): ws3.cell(row=i,column=ci).fill = RED_FILL
        elif k["performanceBand"] in ("Exceptional (A)","Exceeds (B)"):
            for ci in range(1,10): ws3.cell(row=i,column=ci).fill = GREEN_FILL
    auto_width(ws3)

    # Sheet 4: Project_History
    ws4 = wb.create_sheet("Project_History"); add_title(ws4, f"Project History - {name}")
    write_header(ws4, ["projectId","role","contributionSummary","individualOutcome","hasMistake","mistakeIssue","recoveryAction"], 3)
    for i,a in enumerate(assign_by_pk.get(pk,[]),4):
        hm = "Yes" if a.get("hasMistake") else "No"
        write_row(ws4,i,[a["projectId"],a["role"],a["contributionSummary"],a["individualOutcome"],hm,a.get("mistakeIssue",""),a.get("recoveryAction","")])
        if hm=="Yes":
            for ci in range(1,8): ws4.cell(row=i,column=ci).fill = YELLOW_FILL
    auto_width(ws4)

    # Sheet 5: Collaboration_Network
    ws5 = wb.create_sheet("Collaboration_Network"); add_title(ws5, f"Collaboration Network - {name}")
    write_header(ws5, ["collaboratorEmployeeId","collaboratorName","collaboratorDept","projectId","relationshipType","collaborationQuality","hasConflict","conflictSummary","resolutionSummary"], 3)
    seen,deduped = set(),[]
    for c in collab_by_pk.get(pk,[]):
        key = (min(c["employeeId"],c["collaboratorEmployeeId"]),max(c["employeeId"],c["collaboratorEmployeeId"]),c["projectId"])
        if key not in seen: seen.add(key); deduped.append(c)
    for i,c in enumerate(deduped[:50],4):
        op = c["collaboratorEmployeeId"]; o = employees_by_pk.get(op,{})
        hc = "Yes" if c.get("hasConflict") else "No"
        write_row(ws5,i,[op,o.get("name","Unknown"),o.get("department",""),c["projectId"],
            c.get("relationshipTypeName",c.get("relationshipType","")),c["collaborationQuality"],hc,
            c.get("conflictSummary",""),c.get("resolutionSummary","")])
        if hc=="Yes":
            for ci in range(1,10): ws5.cell(row=i,column=ci).fill = YELLOW_FILL
    auto_width(ws5)

    # Sheet 6: Warning_Disciplinary_History
    ws6 = wb.create_sheet("Warning_Disciplinary_History"); add_title(ws6, f"Warning / Disciplinary History - {name}")
    write_header(ws6, ["caseId","caseDate","caseType","severity","formalWarning","summary","rootCause","actionTaken","resolutionStatus","managerInvolved","hrConfidentialityLevel","redactionRequired","linkedProjectId","linkedTrainingId"], 3)
    for i,w in enumerate(sorted(warn_by_pk.get(pk,[]),key=lambda x:x["caseDate"]),4):
        fw = "Yes" if w.get("formalWarning") else "No"; rd = "Yes" if w.get("redactionRequired") else "No"
        write_row(ws6,i,[w["caseId"],w["caseDate"],w["caseType"],w["severity"],fw,w["summary"],w["rootCause"],w["actionTaken"],w["resolutionStatus"],w.get("managerInvolved",""),w.get("hrConfidentialityLevel",""),rd,w.get("linkedProjectId",""),w.get("linkedTrainingId","")])
        if w["severity"] in ("High","Critical"):
            for ci in range(1,15): ws6.cell(row=i,column=ci).fill = RED_FILL
    auto_width(ws6)

    # Sheet 7: Learning_Development
    ws7 = wb.create_sheet("Learning_Development"); add_title(ws7, f"Learning & Development - {name}")
    write_header(ws7, ["trainingId","trainingName","required","completionStatus","completionDate","skillArea","skillLevelBefore","skillLevelAfter","relatedMistake","relatedProjectId","notes"], 3)
    for i,l in enumerate(learn_by_pk.get(pk,[]),4):
        req = "Yes" if l.get("required") else "No"
        write_row(ws7,i,[l["trainingId"],l["trainingName"],req,l["completionStatus"],l.get("completionDate",""),l["skillArea"],l["skillLevelBefore"],l["skillLevelAfter"],l.get("relatedMistake",""),l.get("relatedProjectId",""),l.get("notes","")])
        if l["completionStatus"]!="Completed":
            for ci in range(1,12): ws7.cell(row=i,column=ci).fill = YELLOW_FILL
    auto_width(ws7)

    # Sheet 8: IT_Asset_Register
    ws8 = wb.create_sheet("IT_Asset_Register"); add_title(ws8, f"IT Asset Register - {name}")
    write_header(ws8, ["Asset_Type","Brand","Cost_THB","Status"], 3)
    for i,a in enumerate(it_assets_by_pk.get(pk,[]),4): write_row(ws8,i,[a["Asset_Type"],a["Brand"],a["Cost_THB"],a["Status"]])
    auto_width(ws8)

    # Sheet 9: IT_Ticket_Log
    ws9 = wb.create_sheet("IT_Ticket_Log"); add_title(ws9, f"IT Ticket Log - {name}")
    write_header(ws9, ["Ticket_Issue","Status"], 3)
    for i,t in enumerate(it_tickets_by_pk.get(pk,[]),4): write_row(ws9,i,[t["Ticket_Issue"],t["Status"]])
    auto_width(ws9)

    # Sheet 10: Software_Licenses
    ws10 = wb.create_sheet("Software_Licenses"); add_title(ws10, f"Software Licenses - {name}")
    write_header(ws10, ["License_Name","Status"], 3)
    for i,l in enumerate(sw_licenses_by_pk.get(pk,[]),4): write_row(ws10,i,[l["License_Name"],l["Status"]])
    auto_width(ws10)

    # Sheet 11: Salary_History
    ws11 = wb.create_sheet("Salary_History"); add_title(ws11, f"Salary History - {name}")
    write_header(ws11, ["Base_Salary","Bonus_Months","Increase_Percent"], 3)
    sal = salary_by_pk.get(pk,{})
    if sal: write_row(ws11,4,[sal["Base_Salary"],sal["Bonus_Months"],sal["Increase_Percent"]])
    auto_width(ws11)

    # Sheet 12: Attendance_Record
    ws12 = wb.create_sheet("Attendance_Record"); add_title(ws12, f"Attendance Record - {name}")
    write_header(ws12, ["Sick_Leave_Days","Personal_Leave_Days","Late_Arrivals"], 3)
    att = attendance_by_pk.get(pk,{})
    if att: write_row(ws12,4,[att["Sick_Leave_Days"],att["Personal_Leave_Days"],att["Late_Arrivals"]])
    auto_width(ws12)

    # ── LOOP 22: Sheet 13: 360_Feedback ──
    ws13 = wb.create_sheet("360_Feedback"); add_title(ws13, f"360 Feedback - {name}")
    write_header(ws13, ["Reviewer_Type","Comment"], 3)
    fb = feedback_by_pk.get(pk,{})
    if fb: write_row(ws13,4,[fb["Reviewer_Type"],fb["Comment"]])
    auto_width(ws13)

    # ── LOOP 22: Sheet 14: Skill_Matrix ──
    ws14 = wb.create_sheet("Skill_Matrix"); add_title(ws14, f"Skill Matrix - {name}")
    write_header(ws14, ["Core_Skill","Language_Score_IELTS","Certification"], 3)
    sk = skill_by_pk.get(pk,{})
    if sk: write_row(ws14,4,[sk["Core_Skill"],sk["Language_Score_IELTS"],sk["Certification"]])
    auto_width(ws14)

    # ── LOOP 22: Sheet 15: Succession_Planning ──
    ws15 = wb.create_sheet("Succession_Planning"); add_title(ws15, f"Succession Planning - {name}")
    write_header(ws15, ["Flight_Risk_Pct","Impact_of_Loss","Readiness"], 3)
    sp = succession_by_pk.get(pk,{})
    if sp: write_row(ws15,4,[sp["Flight_Risk_Pct"],sp["Impact_of_Loss"],sp["Readiness"]])
    auto_width(ws15)

    # ── LOOP 22: Sheet 16: Benefit_Claims ──
    ws16 = wb.create_sheet("Benefit_Claims"); add_title(ws16, f"Benefit Claims - {name}")
    write_header(ws16, ["Medical_THB","Dental_THB","Wellness_THB"], 3)
    bc = benefits_by_pk.get(pk,{})
    if bc: write_row(ws16,4,[bc["Medical_THB"],bc["Dental_THB"],bc["Wellness_THB"]])
    auto_width(ws16)

    # ── LOOP 22: Sheet 17: Expense_Reports ──
    ws17 = wb.create_sheet("Expense_Reports"); add_title(ws17, f"Expense Reports - {name}")
    write_header(ws17, ["Travel_THB","Entertainment_THB","Office_Supplies_THB"], 3)
    ex = expenses_by_pk.get(pk,{})
    if ex: write_row(ws17,4,[ex["Travel_THB"],ex["Entertainment_THB"],ex["Office_Supplies_THB"]])
    auto_width(ws17)

    # ── LOOP 22: Sheet 18: Grievance_Log ──
    ws18 = wb.create_sheet("Grievance_Log"); add_title(ws18, f"Grievance Log - {name}")
    write_header(ws18, ["Complaint_Type","Status"], 3)
    for i,g in enumerate(grievances_by_pk.get(pk,[]),4): write_row(ws18,i,[g["Complaint_Type"],g["Status"]])
    auto_width(ws18)

    # ── LOOP 22: Sheet 19: Compliance_Mandates ──
    ws19 = wb.create_sheet("Compliance_Mandates"); add_title(ws19, f"Compliance Mandates - {name}")
    write_header(ws19, ["Mandate","Status"], 3)
    for i,c in enumerate(compliance_by_pk.get(pk,[]),4): write_row(ws19,i,[c["Mandate"],c["Status"]])
    auto_width(ws19)

    # ── LOOP 22: Sheet 20: Onboarding_Journey ──
    ws20 = wb.create_sheet("Onboarding_Journey"); add_title(ws20, f"Onboarding Journey - {name}")
    write_header(ws20, ["Culture_Fit_Score","Interview_Score","Buddy_Name"], 3)
    oj = onboarding_by_pk.get(pk,{})
    if oj: write_row(ws20,4,[oj["Culture_Fit_Score"],oj["Interview_Score"],oj["Buddy_Name"]])
    auto_width(ws20)

    # ── LOOP 22: Sheet 21: Employee_Engagement ──
    ws21 = wb.create_sheet("Employee_Engagement"); add_title(ws21, f"Employee Engagement - {name}")
    write_header(ws21, ["eNPS","Burnout_Risk"], 3)
    eg = engagement_by_pk.get(pk,{})
    if eg: write_row(ws21,4,[eg["eNPS"],eg["Burnout_Risk"]])
    auto_width(ws21)

    # ── LOOP 22: Sheet 22: Physical_Security ──
    ws22 = wb.create_sheet("Physical_Security"); add_title(ws22, f"Physical Security - {name}")
    write_header(ws22, ["Access_Zone","Parking_Slot","Last_Badge_Swipe"], 3)
    ps = security_by_pk.get(pk,{})
    if ps: write_row(ws22,4,[ps["Access_Zone"],ps["Parking_Slot"],ps["Last_Badge_Swipe"]])
    auto_width(ws22)

    # ── LOOP 22: Sheet 23: Timesheet_Log ──
    ws23 = wb.create_sheet("Timesheet_Log"); add_title(ws23, f"Timesheet Log - {name}")
    write_header(ws23, ["Billable_Hours_Pct","Admin_Hours_Pct"], 3)
    ts = timesheet_by_pk.get(pk,{})
    if ts: write_row(ws23,4,[ts["Billable_Hours_Pct"],ts["Admin_Hours_Pct"]])
    auto_width(ws23)

    wb.save(fp)
    return f"{code}_OneDrive_Profile.xlsx"

# =====================================================
# MAIN GENERATION
# =====================================================
print("=" * 60)
print("GENERATING 150 EMPLOYEE EXCEL FILES (23 sheets)")
print("=" * 60)
errors = []
for pk in range(1, 151):
    try:
        gen_employee(pk)
        if pk % 25 == 0: print(f"  {pk}/150...")
    except Exception as e:
        errors.append(f"PK={pk}: {e}")
        print(f"  ERROR PK={pk}: {e}")
print(f"  Done. {150 - len(errors)}/150 generated. Errors: {len(errors)}")

# =====================================================
# MASTER INDEX EXCEL
# =====================================================
print("\nGenerating Master Index Excel...")
mp = OUTPUT_DIR / "BuildersEye_HR_Master_Index.xlsx"
wb = openpyxl.Workbook()

ws1 = wb.active; ws1.title = "Employee_Directory"
add_title(ws1, "BuildersEye HR Master Index", "Demo dataset - 150 employees | Generated: 2026-06-26")
keys = list(master_index[0].keys()) if master_index else []
write_header(ws1, keys, 3)
for i, row in enumerate(master_index, 4): write_row(ws1, i, [row.get(k, "") for k in keys])
auto_width(ws1)

ws2 = wb.create_sheet("Department_Summary"); add_title(ws2, "Department Summary")
dh = ["Department", "Headcount", "Avg KPI", "Total Warnings", "High/Critical Warnings",
      "R1-Critical", "R2-High", "R3-Moderate", "R4-Low"]
write_header(ws2, dh, 3)
dd = defaultdict(lambda: {"c": 0, "ks": 0, "wt": 0, "wh": 0, "r1": 0, "r2": 0, "r3": 0, "r4": 0})
for row in master_index:
    d = row["department"]; dd[d]["c"] += 1; dd[d]["ks"] += row["latestKpiScore"] or 0
    dd[d]["wt"] += row["warningCount"]
    dd[d]["wh"] += 1 if row["highestWarningSeverity"] in ("High", "Critical") else 0
    rl = row["hrRiskLevel"]
    if rl == "R1 - Critical": dd[d]["r1"] += 1
    elif rl == "R2 - High": dd[d]["r2"] += 1
    elif rl == "R3 - Moderate": dd[d]["r3"] += 1
    else: dd[d]["r4"] += 1
for i, (d, v) in enumerate(sorted(dd.items()), 4):
    avg = round(v["ks"] / v["c"], 2) if v["c"] else 0
    write_row(ws2, i, [d, v["c"], avg, v["wt"], v["wh"], v["r1"], v["r2"], v["r3"], v["r4"]])
auto_width(ws2)

ws3 = wb.create_sheet("Cross_Reference_Map"); add_title(ws3, "Cross-Reference Validation Map")
write_header(ws3, ["employeeId", "code", "name", "fileName", "managerCode", "managerExists",
    "projectCount", "warningCount", "learningCount", "collaboratorCount"], 3)
for i, row in enumerate(master_index, 4):
    pk = row["employeeId"]; mc = row["managerId"]
    me = "N/A (CEO)" if pk == 1 else "Yes" if any(e["code"] == mc for e in identity["identities"]) else "No"
    write_row(ws3, i, [pk, employees_by_pk[pk]["code"], row["employeeName"],
        row["fileName"], mc, me, len(assign_by_pk.get(pk, [])),
        len(warn_by_pk.get(pk, [])), len(learn_by_pk.get(pk, [])), len(collab_by_pk.get(pk, []))])
auto_width(ws3); wb.save(mp)
print(f"  Saved: {mp}")

# =====================================================
# VALIDATION
# =====================================================
print("\n" + "=" * 60)
print("VALIDATION")
print("=" * 60)

actual = sorted([f for f in os.listdir(OUTPUT_DIR) if f.endswith(".xlsx") and f != "BuildersEye_HR_Master_Index.xlsx"])
results = []

v1 = len(actual) == 150
results.append(("150 employee files", v1, f"Found {len(actual)}"))
print(f"  {'PASS' if v1 else 'FAIL'}: 150 employee files - {len(actual)} found")

v2 = mp.exists()
results.append(("Master index exists", v2, str(mp)))
print(f"  {'PASS' if v2 else 'FAIL'}: Master index exists")

of_count = 0; ms_count = 0
req_23 = {"Employee_Profile", "Career_Timeline", "KPI_OKR_History", "Project_History",
    "Collaboration_Network", "Warning_Disciplinary_History", "Learning_Development",
    "IT_Asset_Register", "IT_Ticket_Log", "Software_Licenses", "Salary_History", "Attendance_Record",
    "360_Feedback", "Skill_Matrix", "Succession_Planning", "Benefit_Claims", "Expense_Reports",
    "Grievance_Log", "Compliance_Mandates", "Onboarding_Journey", "Employee_Engagement",
    "Physical_Security", "Timesheet_Log"}
for fn in actual:
    try:
        wb2 = openpyxl.load_workbook(OUTPUT_DIR / fn)
        sheets = set(wb2.sheetnames)
        missing = req_23 - sheets
        if missing: ms_count += 1
        wb2.close()
    except Exception as e: of_count += 1

v3a = of_count == 0
results.append(("All files open", v3a, f"{of_count} failures"))
print(f"  {'PASS' if v3a else 'FAIL'}: All files open - {of_count} failures")

v3b = ms_count == 0
results.append(("All 23 sheets present", v3b, f"{ms_count} with missing sheets"))
print(f"  {'PASS' if v3b else 'FAIL'}: All 23 sheets present - {ms_count} missing")

band_counts = Counter(k["performanceBand"] for k in kpi_all)
a_count = band_counts.get("Exceptional (A)", 0); e_count = band_counts.get("Unsatisfactory (E)", 0)
b_count = band_counts.get("Exceeds (B)", 0); c_count = band_counts.get("Meets (C)", 0); d_count = band_counts.get("Below (D)", 0)
kpi_diverse = a_count > 0 and e_count > 0
results.append(("KPI band diversity", kpi_diverse, f"A={a_count}, B={b_count}, C={c_count}, D={d_count}, E={e_count}"))
print(f"  {'PASS' if kpi_diverse else 'FAIL'}: KPI band diversity - A={a_count}, B={b_count}, C={c_count}, D={d_count}, E={e_count}")

no_proj = [pk for pk in range(1, 151) if len(assign_by_pk.get(pk, [])) == 0]
no_kpi = [pk for pk in range(1, 151) if len(kpi_by_pk.get(pk, [])) == 0]
no_warn = [pk for pk in range(1, 151) if len(warn_by_pk.get(pk, [])) == 0]
no_learn = [pk for pk in range(1, 151) if len(learn_by_pk.get(pk, [])) == 0]
no_collab = [pk for pk in range(1, 151) if len(collab_by_pk.get(pk, [])) == 0]
results.append(("All have projects", len(no_proj) == 0, f"Missing: {len(no_proj)}"))
results.append(("All have KPIs", len(no_kpi) == 0, f"Missing: {len(no_kpi)}"))
results.append(("All have warnings", len(no_warn) == 0, f"Missing: {len(no_warn)}"))
results.append(("All have learning", len(no_learn) == 0, f"Missing: {len(no_learn)}"))
results.append(("All have collaborators", len(no_collab) == 0, f"Missing: {len(no_collab)}"))
print(f"  Coverage - Proj:{len(no_proj)} KPI:{len(no_kpi)} Warn:{len(no_warn)} Learn:{len(no_learn)} Collab:{len(no_collab)}")

cp = OUTPUT_DIR / "validation_report.md"
all_ok = all(ok for _, ok, _ in results)
lines = [
    "# BuildersEye HR OneDrive Demo - Validation Report", "",
    f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", "",
    f"**Employee files:** {len(actual)}", f"**Master index:** BuildersEye_HR_Master_Index.xlsx", "",
    "## Validation Results", "", "| # | Check | Result | Detail |", "|---|---|---|---|",
]
for i, (label, ok, detail) in enumerate(results, 1): lines.append(f"| {i} | {label} | {'PASS' if ok else 'FAIL'} | {detail} |")
lines += [
    "", "## Dataset Inventory", "", "| Dataset | Records |", "|---|---|",
    f"| identity-graph.json | {len(identity['identities'])} employees |",
    f"| career-story-plan.json | {len(careers)} stories |",
    f"| kpi-okr-history.json | {len(kpi_all)} review periods |",
    f"| project-assignments.json | {len(assignments_all)} assignments |",
    f"| collaboration-graph.json | {len(collabs_all)} edges |",
    f"| warning-history.json | {len(warnings_all)} cases |",
    f"| learning-history.json | {len(learning_all)} training records |",
    f"| master-index.json | {len(master_index)} employees |",
    f"| it-assets.json | {len(it_assets_all)} assets |",
    f"| it-tickets.json | {len(it_tickets_all)} tickets |",
    f"| software-licenses.json | {len(sw_licenses_all)} licenses |",
    f"| salary-history.json | {len(salary_all)} employees |",
    f"| attendance-record.json | {len(attendance_all)} employees |",
    f"| feedback-360.json | {len(feedback_all)} reviews |",
    f"| skill-matrix.json | {len(skill_all)} skills |",
    f"| succession-planning.json | {len(succession_all)} plans |",
    f"| benefit-claims.json | {len(benefits_all)} claims |",
    f"| expense-reports.json | {len(expenses_all)} reports |",
    f"| grievance-log.json | {len(grievances_all)} grievances |",
    f"| compliance-mandates.json | {len(compliance_all)} mandates |",
    f"| onboarding-journey.json | {len(onboarding_all)} journeys |",
    f"| employee-engagement.json | {len(engagement_all)} surveys |",
    f"| physical-security.json | {len(security_all)} records |",
    f"| timesheet-log.json | {len(timesheet_all)} logs |",
    f"| **Excel files** | **{len(actual)}** employee files + 1 master index (23 sheets each) |",
    "", "## Final Verdict", "",
    f"**Overall: {'ALL CHECKS PASSED' if all_ok else 'SOME CHECKS FAILED'}**",
    "", f"**{sum(1 for _, ok, _ in results if ok)}/{len(results)} checks passed**",
]
cp.write_text("\n".join(lines), encoding="utf-8")
print(f"\nReport: {cp}")

print(f"\n{'='*60}")
print(f"{'ALL CHECKS PASSED' if all_ok else 'SOME CHECKS FAILED'}")
print(f"{sum(1 for _, ok, _ in results if ok)}/{len(results)} checks passed")
print(f"{'='*60}")