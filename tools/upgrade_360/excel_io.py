# tools/upgrade_360/excel_io.py
"""อ่าน/เขียน Excel พนักงาน 23 sheets (engine openpyxl) — เก็บ header/dimension เดิม.

โครงสร้างไฟล์ต้นฉบับ (ยืนยันจากไฟล์จริง EMP001_OneDrive_Profile.xlsx):
    row 1 : title  ("Collaboration Network - ธนกฤต ศรีสุวรรณ")
    row 2 : blank / metadata ("Department: ... | Generated: ...")
    row 3 : header (ชื่อคอลัมน์)
    row 4+ : data rows

หลักการสำคัญ:
- append ต่อท้าย data rows เท่านั้น — ไม่ทับ header, ไม่ทำ header ซ้ำ, ไม่ทำ sheet ซ้ำ
- ถ้าแถวใหม่มีคอลัมน์ที่ header ยังไม่มี (เช่น logDateTime, eventId, counterpartyEmployeeCode
  จาก DESIGN.md) → ขยาย header row ให้อัตโนมัติ แล้วเขียนค่าให้ตรงคอลัมน์
- copy style พื้นฐาน (font/border/fill/number_format) จาก header row ไปให้แถวที่ append
"""
from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Union

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from config import DEFAULT_SHEET_NAMES

# ── ค่าคงที่โครงสร้าง ─────────────────────────────────────────────────────
TITLE_ROW_IDX = 0          # 0-based: แถว title
HEADER_ROW_IDX = 2         # 0-based: แถว header (row 3 ใน Excel)
# detect header row: แถวแรก (0-based) ที่มี non-empty cell >= 2
# (title row มี 1 cell, blank row มี 0 cell, header row มี >=2 cells)

# ── helpers: path ─────────────────────────────────────────────────────────

def ensure_output_dir(path: Union[str, Path]) -> Path:
    p = Path(path)
    p.mkdir(parents=True, exist_ok=True)
    return p


def default_template_for(emp_file: Union[str, Path]) -> Path:
    """คืน path ไฟล์ต้นฉบับที่ควรใช้เป็น template (ถ้า output ยังไม่มีไฟล์)."""
    return Path(emp_file)


# ── helpers: detect โครงสร้าง ─────────────────────────────────────────────

def _find_header_row_in_frame(df: pd.DataFrame) -> int:
    """หาแถว header (0-based) จาก DataFrame ที่อ่านด้วย header=None.

    ใช้กติกา: แถวแรกที่มี non-null >= 2 ค่า
    """
    for i in range(min(len(df), 10)):
        non_null = df.iloc[i].notna().sum()
        if non_null >= 2:
            return i
    return HEADER_ROW_IDX  # fallback ตาม schema เดิม


def _find_header_row_in_ws(ws: Worksheet) -> int:
    """หาแถว header (0-based) จาก openpyxl worksheet."""
    for i, row in enumerate(ws.iter_rows(max_row=min(ws.max_row, 10), max_col=min(ws.max_column, 40))):
        non_null = sum(1 for c in row if c.value is not None and str(c.value).strip() != "")
        if non_null >= 2:
            return i
    return HEADER_ROW_IDX


def sheet_names_of(path: Union[str, Path]) -> List[str]:
    """คืนชื่อ sheets ทั้งหมดในไฟล์ (อ่านด้วย openpyxl เท่านั้น ไม่ต้อง pandas)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


# ── อ่าน (pandas) ─────────────────────────────────────────────────────────

def read_employee_raw(path: Union[str, Path]) -> Dict[str, pd.DataFrame]:
    """อ่านทุก sheet เป็น DataFrame(header=None) — เก็บทุก cell ตามต้นฉบับ.

    ใช้เป็น raw source สำหรับ snapshot/restore หรือ debug
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"ไม่พบไฟล์ Excel: {p}")
    return pd.read_excel(p, sheet_name=None, header=None, engine="openpyxl")


def read_employee(path: Union[str, Path]) -> Dict[str, pd.DataFrame]:
    """อ่าน 23 sheets เป็น dict[str, DataFrame] — header ตามตำแหน่งจริง, data ตามมา.

    แต่ละ DataFrame: columns = header cells, index reset, data rows ต่อจาก header
    """
    raw = read_employee_raw(path)
    out: Dict[str, pd.DataFrame] = {}
    for name, df in raw.items():
        hdr_idx = _find_header_row_in_frame(df)
        header = [str(v) if v is not None else "" for v in df.iloc[hdr_idx].tolist()]
        data = df.iloc[hdr_idx + 1:].copy()
        data.columns = header
        # ตัดแถวว่างท้ายสุดออก (ถ้ามี)
        data = data.dropna(how="all").reset_index(drop=True)
        out[name] = data
    return out


def read_employee_meta(path: Union[str, Path]) -> Dict[str, Dict[str, Any]]:
    """อ่าน dimension + merged cells + column widths ของทุก sheet (ผ่าน openpyxl).

    ใช้เก็บ 'ของเดิม' ไว้ก่อนเขียน เพื่อให้มั่นใจว่าไม่เสียโครงสร้าง
    """
    wb = load_workbook(path, data_only=False)
    meta: Dict[str, Dict[str, Any]] = {}
    for ws in wb.worksheets:
        meta[ws.title] = {
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "dimensions": ws.dimensions,
            "merged": [str(r) for r in ws.merged_cells.ranges],
            "col_widths": {
                col: ws.column_dimensions[col].width
                for col in ws.column_dimensions if ws.column_dimensions[col].width
            },
            "freeze": str(ws.freeze_panes) if ws.freeze_panes else None,
        }
    wb.close()
    return meta


# ── เขียน / append (openpyxl — รักษา format เดิม) ─────────────────────────

def _copy_cell_style(src: Cell, dst: Cell) -> None:
    """copy style พื้นฐานจาก src ไป dst (font/border/fill/alignment/number_format)."""
    try:
        if src.has_style:
            dst.font = Font(
                name=src.font.name, size=src.font.size, bold=src.font.bold,
                italic=src.font.italic, color=src.font.color,
            )
            dst.border = Border(
                left=src.border.left, right=src.border.right,
                top=src.border.top, bottom=src.border.bottom,
            )
            dst.fill = PatternFill(
                fill_type=src.fill.fill_type,
                fgColor=src.fill.fgColor, bgColor=src.fill.bgColor,
            )
            dst.alignment = Alignment(
                horizontal=src.alignment.horizontal,
                vertical=src.alignment.vertical,
                wrap_text=src.alignment.wrap_text,
            )
            dst.number_format = src.number_format
    except Exception:
        pass  # style ไม่ critical — ข้อมูลสำคัญกว่า


def _ensure_header_columns(ws: Worksheet, header_row_1based: int, col_names: Sequence[str]) -> Dict[str, int]:
    """ขยาย header row ให้มีคอลัมน์ครบตาม col_names (เพิ่มเฉพาะที่ยังไม่มี).

    คืน dict: {col_name: 0-based column index}
    """
    existing: Dict[str, int] = {}
    for j, cell in enumerate(ws[header_row_1based]):
        if cell.value is not None and str(cell.value).strip() != "":
            existing[str(cell.value).strip()] = j
    for name in col_names:
        if name not in existing:
            j = ws.max_column + 1  # 1-based คอลัมน์ว่างถัดไป (ไม่ทับคอลัมน์เดิม)
            hdr_cell = ws.cell(row=header_row_1based, column=j, value=name)
            _copy_cell_style(ws.cell(row=header_row_1based, column=1), hdr_cell)
            existing[name] = j - 1
    return existing


def _rows_to_records(rows: Union[pd.DataFrame, Iterable[Sequence[Any]], Iterable[Dict[str, Any]]]) -> List[Dict[str, Any]]:
    """แปลง input rows ให้เป็น list of dict (ชื่อคอลัมน์ -> ค่า)."""
    if isinstance(rows, pd.DataFrame):
        cols = [str(c) for c in rows.columns]
        return [dict(zip(cols, [r[c] for c in rows.columns])) for _, r in rows.iterrows()]
    out: List[Dict[str, Any]] = []
    for r in rows:
        if isinstance(r, dict):
            out.append({str(k): v for k, v in r.items()})
        else:
            # positional list/tuple -> คอลัมน์ "0","1","2"... ต้อง map หลังรู้ header
            out.append({str(i): v for i, v in enumerate(r)})
    return out


def append_rows(ws: Worksheet, rows: Union[pd.DataFrame, Iterable[Any]]) -> int:
    """append rows ต่อท้าย data rows ของ sheet (ไม่ทับ header, ไม่ทำ header ซ้ำ).

    - รองรับ DataFrame (align ตามชื่อคอลัมน์) / list of dict / list of list
    - คอลัมน์ใหม่ที่ header ยังไม่มี → เพิ่ม header cell ให้อัตโนมัติ
    - copy style จาก header row ไปยังแถวใหม่
    คืนจำนวนแถวที่ append
    """
    hdr_idx = _find_header_row_in_ws(ws)
    header_row_1based = hdr_idx + 1
    first_data_row_1based = hdr_idx + 2

    records = _rows_to_records(rows)
    if not records:
        return 0

    col_names = list(dict.fromkeys(k for rec in records for k in rec.keys()))
    col_index = _ensure_header_columns(ws, header_row_1based, col_names)

    # เริ่มต้นต่อจากแถวข้อมูลสุดท้าย (แต่ไม่ต่ำกว่า first_data_row)
    start = max(ws.max_row + 1, first_data_row_1based)
    hdr_cells = list(ws[header_row_1based])

    for i, rec in enumerate(records):
        r = start + i
        for name, val in rec.items():
            j = col_index.get(name)
            if j is None:
                continue
            col = j + 1
            cell = ws.cell(row=r, column=col, value=val)
            # copy style จาก header cell ถ้ามี
            if j < len(hdr_cells):
                _copy_cell_style(hdr_cells[j], cell)
    return len(records)



def write_employee(
    path: Union[str, Path],
    sheets: Optional[Dict[str, pd.DataFrame]] = None,
    append_rows_per_sheet: Optional[Dict[str, Union[pd.DataFrame, Iterable[Any]]]] = None,
    template_path: Optional[Union[str, Path]] = None,
) -> Path:
    """เขียน/append Excel พนักงาน 1 คน (23 sheets) — ไม่ทำ header/sheet ซ้ำ.

    Args:
        path: ไฟล์ output (จะถูกสร้าง/เขียนทับ)
        sheets: ถ้าให้มา — เขียนใหม่ทั้ง workbook จาก dict of DataFrames
                (header ใช้ df.columns, data = df rows) ใช้ตอน regenerate ไฟล์เต็ม
        append_rows_per_sheet: ถ้าให้มา — append ต่อท้าย data rows เดิมของแต่ละ sheet
                (ใช้ตอน incremental ต่อคน / ต่อ event)
        template_path: ไฟล์ต้นฉบับที่ใช้ copy เป็น base ถ้า path ยังไม่มี
                (default: path เอง — ถ้าไม่มีไฟล์จะ error)

    Returns:
        Path ที่เขียนสำเร็จ
    """
    dst = Path(path)
    ensure_output_dir(dst.parent)

    if dst.exists():
        wb = load_workbook(dst)  # รักษา format เดิม
    else:
        tmpl = Path(template_path) if template_path else Path(path)
        if tmpl.exists():
            shutil.copyfile(tmpl, dst)
            wb = load_workbook(dst)
        else:
            # สร้าง workbook ใหม่ (sheet ตาม sheets/append ที่ให้มา)
            wb = Workbook()
            wb.remove(wb.active)

    try:
        # ── โหมด 1: เขียนใหม่ทั้ง workbook ──
        if sheets is not None:
            for name, df in sheets.items():
                if name in wb.sheetnames:
                    ws = wb[name]
                    # clear เฉพาะแถวข้อมูล (หลัง header) — เก็บ title/header เดิม
                    hdr_idx = _find_header_row_in_ws(ws)
                    if ws.max_row > hdr_idx + 1:
                        ws.delete_rows(hdr_idx + 2, ws.max_row - (hdr_idx + 1))
                    header_1based = hdr_idx + 1
                else:
                    ws = wb.create_sheet(title=name)
                    header_1based = 1  # sheet ใหม่ — header แถวแรก

                if header_1based == 1 and ws.max_row < 1:
                    # เขียน header ลงแถวแรกของ sheet ใหม่
                    for j, col in enumerate(df.columns, start=1):
                        ws.cell(row=1, column=j, value=str(col))
                else:
                    # ตรวจว่า header เดิมตรงกับ df.columns ไหม — ถ้าไม่ตรง ให้ขยาย
                    _ensure_header_columns(ws, header_1based, [str(c) for c in df.columns])

                # เขียน data rows ต่อจาก header
                data = df.dropna(how="all")
                if not data.empty:
                    _append_rows_with_df(ws, df, header_row_1based=header_1based)

        # ── โหมด 2: append ต่อท้าย ──
        if append_rows_per_sheet:
            for name, rows in append_rows_per_sheet.items():
                if name not in wb.sheetnames:
                    wb.create_sheet(title=name)
                ws = wb[name]
                append_rows(ws, rows)

        wb.save(dst)
    finally:
        wb.close()
    return dst


def _append_rows_with_df(ws: Worksheet, df: pd.DataFrame, header_row_1based: int) -> int:
    """เขียน DataFrame ทั้งก้อนต่อจาก header (align ตามชื่อคอลัมน์)."""
    col_index = _ensure_header_columns(ws, header_row_1based, [str(c) for c in df.columns])
    start = max(ws.max_row + 1, header_row_1based + 1)
    hdr_cells = list(ws[header_row_1based])
    written = 0
    for _, rec in df.iterrows():
        r = start + written
        for name, val in rec.items():
            j = col_index.get(str(name))
            if j is None:
                continue
            cell = ws.cell(row=r, column=j + 1, value=val)
            if j < len(hdr_cells):
                _copy_cell_style(hdr_cells[j], cell)
        written += 1
    return written


# ── ของพ่วงท้ายที่ phase 2/3 ใช้บ่อย ──────────────────────────────────────

def list_employee_files(input_dir: Union[str, Path]) -> List[Path]:
    """คืน list ไฟล์ EMP###_OneDrive_Profile.xlsx เรียงตาม code (EMP001..EMP150)."""
    d = Path(input_dir)
    files = sorted(
        d.glob("EMP*_OneDrive_Profile.xlsx"),
        key=lambda p: p.name,
    )
    return files


def emp_code_from_file(path: Union[str, Path]) -> str:
    """EMP001_OneDrive_Profile.xlsx -> 'EMP001'"""
    return Path(path).name.split("_")[0]


if __name__ == "__main__":
    # smoke test: python excel_io.py <path-to-emp-file>
    import sys

    p = sys.argv[1] if len(sys.argv) > 1 else "../../src/data/hr_onedrive_demo/EMP001_OneDrive_Profile.xlsx"
    sheets = read_employee(p)
    print(f"อ่าน {Path(p).name}: {len(sheets)} sheets")
    for name, df in sheets.items():
        print(f"  {name}: {df.shape}")

