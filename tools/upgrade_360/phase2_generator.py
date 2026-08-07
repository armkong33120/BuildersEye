# -*- coding: utf-8 -*-
"""phase2_generator.py — Phase 2 Data Generation Engine (หัวใจ)

ฟังก์ชันหลัก: :func:`generate_employee(emp, ctx)` สำหรับพนักงาน 1 คน

ขั้นตอนต่อคน:
1. อ่าน ``relationship_matrix.json`` + ``storyline_catalog.json`` จาก ``design/``
2. อ่าน Excel ของคนนั้น — ผ่าน ``excel_io.read_employee`` (ของ core-dev, import lenient)
   ถ้ายังไม่มี module นั้น → fallback ด้วย openpyxl (ฟังก์ชันภายในไฟล์นี้)
3. เลือก events ที่คนนี้เป็น participant (suggestedParticipants / keyPairCodes /
   คู่ใน matrix ที่อ้าง eventId) → สร้าง log แทรกลง affectedSheets
   (Grievance_Log, Warning_Disciplinary_History, Collaboration_Network,
    Expense_Reports, Project_History, Attendance_Record, Timesheet_Log ฯลฯ)
4. เติม routine logs จนได้ 300-500 rows/คน (80% routine / 20% drama)
5. ทุก drama event คู่คน: eventId + logDateTime เดียวกันทั้ง 2 ฝั่ง
   (ผ่าน ``ctx.pair_registry`` + ``ctx.event_times``) และบันทึก metadata ลง
   ``ctx.injected_events`` (list) → Phase 3 ใช้ตรวจข้ามไฟล์
6. ไม่ยิง API ซ้ำ event เดิม (dedupe ผ่าน ``ctx.drama_descriptions`` / ``drama_emitted``)
7. ปริมาณดราม่าตามแผนก/ความอาวุโส (senior ได้ drama เยอะกว่า)

ctx: :class:`DataGenContext` — config (no_api/api_key/workers), injected_events
global (idempotent), progress log + checkpoint (re-run ไม่ซ้ำ)

CLI:
  python phase2_generator.py --codes EMP005,EMP101 --no-api
  python phase2_generator.py --all --no-api --workers 4
  python phase2_generator.py --codes EMP005 --api-key $DEEPSEEK_API_KEY

เจ้าของไฟล์: Data Generation (Phase 2)
"""

from __future__ import annotations

import argparse
import collections
import json
import os
import random
import threading
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import openpyxl

try:  # รันแบบ package
    from .pydantic_models import COMMON_COLS, DramaEventInjection, RoutineLogRow, TimelineRow
    from .deepseek_client import DeepSeekDramaClient
    from .faker_routines import (PROJECT_IDS, SHEET_ORIGINAL_COLS, SHEET_SOURCE,
                                 generate_routine_plan, orig_values_for)

except ImportError:  # รันแบบ script ตรง
    from pydantic_models import COMMON_COLS, DramaEventInjection, RoutineLogRow, TimelineRow
    from deepseek_client import DeepSeekDramaClient
    from faker_routines import (PROJECT_IDS, SHEET_ORIGINAL_COLS, SHEET_SOURCE,
                                generate_routine_plan, orig_values_for)

# ---------------------------------------------------------------------------
# Paths / Constants
# ---------------------------------------------------------------------------
TOOLS_DIR = Path(__file__).resolve().parent            # tools/upgrade_360
APP_ROOT = TOOLS_DIR.parents[1]                        # รากโปรเจกต์
DESIGN_DIR = TOOLS_DIR / "design"
SRC_DATA_DIR = APP_ROOT / "src" / "data"
HR_DIR = SRC_DATA_DIR / "hr_onedrive_demo"
OUTPUT_DIR = TOOLS_DIR / "output" / "hr_onedrive_upgraded"
CHECKPOINT_DIR = TOOLS_DIR / "output"

VALID_SHEETS = frozenset({
    "Employee_Profile", "Career_Timeline", "KPI_OKR_History", "Project_History",
    "Collaboration_Network", "Warning_Disciplinary_History", "Learning_Development",
    "IT_Asset_Register", "IT_Ticket_Log", "Software_Licenses", "Salary_History",
    "Attendance_Record", "360_Feedback", "Skill_Matrix", "Succession_Planning",
    "Benefit_Claims", "Expense_Reports", "Grievance_Log", "Compliance_Mandates",
    "Onboarding_Journey", "Employee_Engagement", "Physical_Security", "Timesheet_Log",
})

# architect เขียน affectedSheets หลุดมา 2 ชื่อ (แผนก ไม่ใช่ sheet) → map ไป sheet จริง
SHEET_ALIAS = {"HR & Admin": "Grievance_Log", "Sales": "Grievance_Log"}

# logType ตาม category (DESIGN.md §5.1 enum)
LOG_TYPE_BY_CATEGORY = {
    "crisis": "legacy_context",
    "politics": "incident",
    "grey_area_collusion": "expense_irregularity",
    "cross_dept_conflict": "incident",
    "dept_negative": "warning",
    "positive": "praise",
    "routine": "routine",
    "family": "incident",
}

REL_NAME_TH = {
    "conflict": "ขัดแย้ง", "collusion": "สมคบ/ทุจริตร่วม", "friendship": "เพื่อน/สนิท",
    "mentorship": "พี่เลี้ยง", "work_partner": "คู่ทำงาน", "family": "ญาติ/ครอบครัว",
}

DEFAULT_CONFIG: Dict[str, Any] = {
    "no_api": True,
    "api_key": "",
    "workers": 1,
    "seed": 20260807,
    "as_of": "2026-08-07",
    "target_rows": 420,
    "min_rows": 300,
    "max_rows": 500,
    "drama_ratio": 0.20,
    "routine_pair_ratio": 0.20,
    "desc_batch_size": 8,   # จำนวน events ต่อ 1 API call (batch ลด token overhead)
    "src_dir": str(HR_DIR),
    "output_dir": str(OUTPUT_DIR),
    "force": False,
}


def _fmt_iso(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%dT%H:%M:%S+07:00")


def _pair_key(a: str, b: str) -> str:
    return "|".join(sorted([a, b]))

# ---------------------------------------------------------------------------
# DataGenContext — config + registries + idempotency + progress
# ---------------------------------------------------------------------------
class DataGenContext:
    """Context กลางของ Phase 2 (โครงสร้างที่ Phase 3 จะอ่านต่อ)

    ฟิลด์ที่ Phase 3 ใช้ตรวจข้ามไฟล์:
      - injected_events : list[dict] — metadata ของทุกแถวที่ inject เข้าไฟล์
        (eventId, logDateTime, sheet, employeeCode, counterpartyEmployeeCode,
         subject, descriptionTH, riskLevel, category, logType, relationship,
         faction, mirrorRequired, mirrorEmployeeCode)
      - event_times     : dict[eventId -> list[iso]] — canonical timestamp ต่อ event
      - pair_registry   : dict[key -> record] — คู่คนที่ต้องมีแถวครบ 2 ฝั่ง
      - progress_log    : list[dict] — สรุปต่อคน (เขียน output/progress.jsonl)
    """

    def __init__(self, config: Optional[Dict[str, Any]] = None):
        cfg = dict(DEFAULT_CONFIG)
        if config:
            cfg.update({k: v for k, v in config.items() if v is not None})
        self.config = cfg
        self.rng = random.Random(int(cfg["seed"]))
        self.lock = threading.RLock()
        self.rng_local = threading.local()

        # ---- ข้อมูลอ้างอิง (โหลดครั้งเดียว) ----
        self.identity = self._load_json(APP_ROOT / "src" / "data" / "identity-graph.json")
        identities = self.identity.get("identities", [])
        self.identity_by_code: Dict[str, Dict[str, Any]] = {e["code"]: e for e in identities}
        self.identity_by_pk: Dict[int, Dict[str, Any]] = {e["pk"]: e for e in identities}

        # hire date ต่อคน (career-story-plan ใช้ employeeId = pk)
        careers = self._load_json(APP_ROOT / "src" / "data" / "career-story-plan.json")
        self.join_date_by_code: Dict[str, date] = {}
        for c in careers:
            emp = self.identity_by_pk.get(c.get("employeeId"))
            if emp and c.get("joinDate"):
                self.join_date_by_code[emp["code"]] = date.fromisoformat(str(c["joinDate"])[:10])

        # ---- design files (architect) ----
        self.matrix = self._load_json(DESIGN_DIR / "relationship_matrix.json")
        self.faction_by_code: Dict[str, str] = dict(self.matrix.get("employeeFaction", {}))
        self.pairs_by_code: Dict[str, List[Dict[str, Any]]] = collections.defaultdict(list)
        for p in self.matrix.get("pairs", []):
            self.pairs_by_code[p["a"]].append(p)
            self.pairs_by_code[p["b"]].append(p)

        cat = self._load_json(DESIGN_DIR / "storyline_catalog.json")
        self.catalog: List[Dict[str, Any]] = cat.get("catalog", [])
        self.event_index: Dict[str, Dict[str, Any]] = {e["eventId"]: e for e in self.catalog}
        self.events_by_code: Dict[str, Set[str]] = collections.defaultdict(set)
        self._build_involvement()

        # ---- client (API / offline) ----
        self.client = DeepSeekDramaClient(
            api_key=cfg.get("api_key") or None,
            no_api=bool(cfg.get("no_api", True)),
            rng=self.rng,
            cost_output_dir=cfg.get("output_dir") or None,
        )

        # ---- registries (idempotent ข้าม run ผ่าน checkpoint) ----
        self.event_times: Dict[str, List[str]] = {}
        self.drama_descriptions: Dict[str, Dict[str, Dict[str, str]]] = {}
        self.injected_events: List[Dict[str, Any]] = []
        self.pair_registry: Dict[str, Dict[str, Any]] = {}
        self.pair_emitted: Set[str] = set()
        self.drama_emitted: Set[str] = set()
        # ---- drama pair-centric (QA fix): occurrence registry + ฝั่งที่ emit แล้ว ----
        # drama_pending      : occ_key (eventId|sheet|t) -> canonical occurrence record
        #                      (participants = ชุด participant เดียวของ event ณ t — ทุกคนใช้ชุดนี้)
        # drama_side_emitted : f"{code}|{occ_key}" — ฝั่งที่สร้างแถวแล้ว (dedupe/กันซ้ำ)
        # drama_participant_cache : f"{eventId}|{t}" -> participants (คำนวณครั้งเดียว)
        self.drama_pending: Dict[str, Dict[str, Any]] = {}
        self.drama_side_emitted: Set[str] = set()
        self.drama_participant_cache: Dict[str, List[str]] = {}
        # _desc_requested: (eventId|sheet) ที่ขอ description แล้ว (กัน duplicate batch ระหว่าง thread)
        self._desc_requested: Set[str] = set()
        # file_locks: serialize การเขียนไฟล์คนเดียวกันเมื่อ workers > 1 (generate + backfill)
        self.file_locks: Dict[str, threading.Lock] = {}
        self._generating: Set[str] = set()   # คนที่กำลัง generate — backfill จะข้าม (คนนั้นจัดการเอง)
        self.progress_log: List[Dict[str, Any]] = []
        self.stats: Dict[str, int] = collections.Counter()
        self._load_checkpoint()

    # ------------------------------------------------------------------
    def get_rng(self) -> random.Random:
        """Random ต่อ thread — ปลอดภัยเมื่อ --workers > 1"""
        r = getattr(self.rng_local, "r", None)
        if r is None:
            r = random.Random(int(self.config["seed"]) + (threading.get_ident() % 100000))
            self.rng_local.r = r
        return r

    def cost_summary(self) -> Dict[str, Any]:
        """สรุปค่าใช้จ่าย API (tokens จริงจาก CostTracker ของ client)"""
        try:
            return self.client.cost_tracker.summary()
        except Exception:
            return {"calls": 0, "estimated_cost_usd": 0.0}

    @staticmethod
    def _load_json(path: Path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)

    def _build_involvement(self) -> None:
        """events_by_code: ทุก event ที่คนนี้เป็น participant
        = suggestedParticipants + keyPairCodes + คู่ใน matrix ที่อ้าง eventId นั้น
        """
        for ev in self.catalog:
            eid = ev["eventId"]
            codes: Set[str] = set()
            codes.update(ev.get("suggestedParticipants", []))
            for a, b in ev.get("keyPairCodes", []):
                codes.add(a)
                codes.add(b)
            for code in codes:
                self.events_by_code[code].add(eid)
        # คู่ใน matrix ที่อ้าง eventId
        for p in self.matrix.get("pairs", []):
            for eid in p.get("eventIds", []):
                self.events_by_code[p["a"]].add(eid)
                self.events_by_code[p["b"]].add(eid)

    def hire_date(self, code: str) -> Optional[date]:
        return self.join_date_by_code.get(code)

    def seniority_factor(self, emp: Dict[str, Any]) -> float:
        """senior ได้ drama เยอะกว่า — depth ต่ำ + บทบาทบริหาร + ขั้วการเมือง"""
        code = emp.get("code", "")
        depth = int(emp.get("hierarchyDepth", 3) or 3)
        f = 1.0 / (depth + 1)
        if emp.get("roleGroup") in ("CEO", "Director", "Manager", "Head", "Team Lead", "Supervisor"):
            f += 0.5
        if self.faction_by_code.get(code) in ("old_guard", "new_guard"):
            f += 0.15
        return min(1.0, f)

    def drama_cap(self, emp: Dict[str, Any]) -> int:
        """จำนวนแถวดราม่าสูงสุดต่อคน (ตามอาวุโส) — clamp 12..30% ของ target"""
        target = int(self.config["target_rows"])
        base = target * float(self.config["drama_ratio"])
        cap = base * (0.5 + self.seniority_factor(emp))
        lo = max(12, int(base * 0.55))
        hi = int(target * 0.30)
        return max(lo, min(hi, int(cap)))

    def events_for(self, code: str) -> List[Dict[str, Any]]:
        """events ของคนนี้ เรียงตามความสำคัญ: primary (suggested/keyPair) ก่อน,
        ตามด้วย riskLevel สูงก่อน, expansion มากก่อน"""
        evs = [self.event_index[e] for e in self.events_by_code.get(code, set()) if e in self.event_index]
        risk_rank = {"critical": 0, "high": 1, "medium": 2, "low": 3}

        def pri(e):
            primary = (code in e.get("suggestedParticipants", []) or
                       any(code in (a, b) for a, b in e.get("keyPairCodes", [])))
            return (0 if primary else 1,
                    risk_rank.get(e.get("riskLevel", "medium"), 4),
                    -int(e.get("logRowExpansion", 1)),
                    e["eventId"])  # tiebreaker กำหนดเอง → deterministic ข้าม run

        evs.sort(key=pri)
        return evs

    # ------------------------------------------------------------------
    # Timestamp แบบ canonical ต่อ event (ทั้ง 2 ฝั่งใช้ค่าเดียวกัน)
    # ------------------------------------------------------------------
    def ensure_event_times(self, ev: Dict[str, Any]) -> List[str]:
        """คืนรายการ timestamp (ISO) ของ event — คำนวณครั้งเดียวต่อ eventId."""
        eid = ev["eventId"]
        with self.lock:
            if eid in self.event_times:
                return self.event_times[eid]
            period = ev.get("period", "ongoing")
            n = 2 if ev.get("recurring") else 1  # recurring → 2 ครั้ง (ควบคุมปริมาณ)
            times: List[str] = []
            for i in range(n):
                times.append(_fmt_iso(self._pick_event_dt(period, i)))
            self.event_times[eid] = times
            return times

    def _pick_event_dt(self, period: str, idx: int) -> datetime:
        """เลือกวัน-เวลาของ event ตาม period (1997/2011/2020/ปีอื่น/ongoing)."""
        as_of = self.config.get("as_of", "2026-08-07")
        end = date.fromisoformat(str(as_of))
        if period == "1997":
            start = date(1997, 5, 1); end = date(1997, 12, 31)
        elif period == "2011":
            start = date(2011, 7, 1); end = date(2011, 12, 31)
        elif period == "2020":
            start = date(2020, 1, 1); end = date(2020, 12, 31)
        elif period.isdigit() and len(period) == 4:
            y = int(period)
            start = date(y, 1, 1); end = date(y, 12, 31)
        else:  # ongoing → 2023 จนถึง as_of
            start = date(2023, 1, 1)
        span = max(1, (end - start).days)
        d = start + timedelta(days=self.get_rng().randint(0, span))
        hour = self.get_rng().choice([9, 10, 11, 13, 14, 15, 16])
        return datetime(d.year, d.month, d.day, hour, self.get_rng().randint(0, 59))

    # ------------------------------------------------------------------
    # คำอธิบายดราม่า — generate ครั้งเดียวต่อ (eventId, sheet) ใช้ร่วมทั้ง 2 ฝั่ง
    # ------------------------------------------------------------------
    def _build_desc_spec(self, ev: Dict[str, Any], sheet: str,
                         spec_extra: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """สร้าง spec คำขอ description (ใช้ร่วมกันระหว่าง prewarm/batch และ fallback)."""
        spec: Dict[str, Any] = {
            "eventId": ev["eventId"],
            "titleTH": ev.get("titleTH", ""),
            "descriptionTH": ev.get("descriptionTH", ""),
            "category": ev.get("category", ""),
            "riskLevel": ev.get("riskLevel", "medium"),
            "sheet": sheet,
            "logType": LOG_TYPE_BY_CATEGORY.get(ev.get("category", ""), "incident"),
            "source": SHEET_SOURCE.get(sheet, "HRIS"),
            "location": self._event_location(ev),
            "logDateTime": self.ensure_event_times(ev)[0],
            "resolutionStatus": ev.get("resolutionStatus", ""),
            "financialImpactTHB": ev.get("financialImpactTHB"),
            "expansion": int(ev.get("logRowExpansion", 1)),
        }
        if spec_extra:
            spec.update({k: v for k, v in spec_extra.items() if v})
        return spec

    def prewarm_descriptions(self, emp: Dict[str, Any]) -> int:
        """Batch-fetch คำอธิบาย (eventId, sheet) ที่คนนี้ต้องใช้ → 1 API call ต่อ N events.

        - วางแผนล่วงหน้า (ก่อน _plan_drama_rows/_plan_collab_pairs/_plan_catalog_routine_rows)
          เพื่อให้ planning เจอ cache หมด → ไม่มี per-event call เหลือ
        - offline/no-api: generate_drama_events fallback template ต่อ spec (ผลเหมือนเดิม)
        - cache ต่อ (eventId, sheet) → ข้ามของที่มีแล้ว (ข้าม run ด้วย checkpoint)
        - _desc_requested ป้องกัน duplicate ระหว่าง thread (workers > 1)
        """
        code = emp["code"]
        specs: List[Dict[str, Any]] = []
        seen: Set[str] = set()

        def add(ev: Dict[str, Any], sheet: str, extra: Dict[str, Any]) -> None:
            eid = ev["eventId"]
            key = f"{eid}|{sheet}"
            with self.lock:
                if key in seen or key in self._desc_requested:
                    return
                if eid in self.drama_descriptions and sheet in self.drama_descriptions[eid]:
                    return
                seen.add(key)
                self._desc_requested.add(key)
            specs.append(self._build_desc_spec(ev, sheet, extra))

        # 1) drama rows — non-routine events → affectedSheets (ยกเว้น Collaboration_Network)
        for ev in self.events_for(code):
            if ev.get("category") == "routine":
                continue
            sheets = [SHEET_ALIAS.get(s, s) for s in ev.get("affectedSheets", [])]
            # counterparty ชุดเดียวกับที่ _plan_drama_rows ส่งให้ prompt (ให้ description มีชื่อคน)
            participants = self.drama_participant_set(ev, self.ensure_event_times(ev)[0])
            counterparties = [p for p in participants if p != code]
            extra = {
                "employeeCode": code, "employeeName": emp.get("name", ""),
                "counterpartyEmployeeCode": ";".join(counterparties[:6]),
                "counterpartyNames": ";".join(
                    self.identity_by_code.get(c, {}).get("name", c) for c in counterparties[:6]),
            }
            for sheet in sheets:
                if sheet in VALID_SHEETS and sheet != "Collaboration_Network":
                    add(ev, sheet, extra)

        # 2) collab pairs — first event ของ pair → Collaboration_Network (mirror 2 ฝั่ง)
        for pair in self.pairs_by_code.get(code, []):
            other = pair["b"] if pair["a"] == code else pair["a"]
            if other not in self.identity_by_code:
                continue
            eids = [e for e in pair.get("eventIds", []) if e in self.event_index]
            if not eids:
                continue
            add(self.event_index[eids[0]], "Collaboration_Network", {
                "employeeCode": code, "employeeName": emp.get("name", ""),
                "counterpartyEmployeeCode": other,
                "counterpartyNames": self.identity_by_code.get(other, {}).get("name", other),
            })

        # 3) catalog routine rows — routine events ที่ code เป็น suggested participant
        for ev in self.events_for(code):
            if ev.get("category") != "routine" or code not in ev.get("suggestedParticipants", []):
                continue
            sheets = [SHEET_ALIAS.get(s, s) for s in ev.get("affectedSheets", [])]
            for sheet in sheets:
                if sheet in VALID_SHEETS and sheet != "Collaboration_Network":
                    add(ev, sheet, {"employeeCode": code, "employeeName": emp.get("name", "")})

        if not specs:
            return 0
        batch = max(1, int(self.config.get("desc_batch_size", 8)))
        fetched = 0
        for i in range(0, len(specs), batch):
            chunk = specs[i:i + batch]
            outs = self.client.generate_drama_events(chunk)
            for spec, out in zip(chunk, outs):
                eid, sheet = spec["eventId"], spec["sheet"]
                with self.lock:
                    self.drama_descriptions.setdefault(eid, {})[sheet] = {
                        "subject": out.get("subject", ""),
                        "descriptionTH": out.get("descriptionTH", ""),
                    }
                fetched += 1
        return fetched

    def ensure_drama_description(self, ev: Dict[str, Any], sheet: str,
                                 spec_extra: Optional[Dict[str, Any]] = None) -> Dict[str, str]:
        eid = ev["eventId"]
        with self.lock:
            if eid in self.drama_descriptions and sheet in self.drama_descriptions[eid]:
                return self.drama_descriptions[eid][sheet]
            spec = self._build_desc_spec(ev, sheet, spec_extra)
            # safety net: ถ้า prewarm ไม่ทัน (เหตุการณ์ใหม่/edge case) → fallback single call
            out = self.client.generate_drama_event(spec)
            self.drama_descriptions.setdefault(eid, {})[sheet] = {
                "subject": out.get("subject", ""),
                "descriptionTH": out.get("descriptionTH", ""),
            }
            return self.drama_descriptions[eid][sheet]

    def _event_location(self, ev: Dict[str, Any]) -> str:
        """สถานที่ของ event — เลือกจาก keyword ในเรื่อง ไม่งั้นสุ่มจาก site/HQ."""
        text = (ev.get("titleTH", "") + " " + ev.get("descriptionTH", ""))
        if "น้ำท่วม" in text or "ไซต์" in text or "หน้างาน" in text:
            return self.get_rng().choice(["site-nonthaburi", "site-bangna", "site-chaengwattana", "site-rangsit"])
        if "ประมูล" in text or "จัดซื้อ" in text or "ซัพพลาย" in text:
            return "HQ"
        if "CRM" in text or "ข้อมูลลูกค้า" in text or "ระบบ" in text:
            return "HQ"
        return self.get_rng().choice(["HQ", "site-nonthaburi", "site-bangna", "site-chaengwattana"])

    # ------------------------------------------------------------------
    # อ่าน/เขียน Excel — excel_io ของ core-dev (import lenient) + fallback
    # ------------------------------------------------------------------
    def read_employee(self, code: str) -> Dict[str, Dict[str, Any]]:
        """คืน dict {sheet: {'headers': [...], 'rows': [[...], ...]}}
        พยายามใช้ excel_io.read_employee (core-dev) ก่อน ถ้าไม่มี/ผิดพลาด → fallback openpyxl.
        """
        path = Path(self.config["src_dir"]) / f"{code}_OneDrive_Profile.xlsx"
        try:
            self._ensure_core_on_path()
            import excel_io  # ของ core-dev (module แบน ใน tools/upgrade_360)
            if hasattr(excel_io, "read_employee"):
                book = excel_io.read_employee(path)
                return self._normalize_core_book(book)
        except Exception:
            pass
        return self._fallback_read(code)

    @staticmethod
    def _ensure_core_on_path() -> None:
        """core-dev เขียน module แบน (excel_io/config/...) → ต้องมี tools/upgrade_360 ใน sys.path"""
        import sys
        if str(TOOLS_DIR) not in sys.path:
            sys.path.insert(0, str(TOOLS_DIR))

    @staticmethod
    def _normalize_core_book(book: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
        """แปลง dict[str, DataFrame] ของ core-dev → {sheet: {headers, rows}}"""
        out: Dict[str, Dict[str, Any]] = {}
        for name, df in book.items():
            headers = [str(c) if c is not None else "" for c in df.columns]
            rows = [df.iloc[i].tolist() for i in range(len(df))]
            out[str(name)] = {"headers": headers, "rows": rows}
        return out

    def _fallback_read(self, code: str) -> Dict[str, Dict[str, Any]]:
        path = Path(self.config["src_dir"]) / f"{code}_OneDrive_Profile.xlsx"
        wb = openpyxl.load_workbook(path, data_only=True)
        book: Dict[str, Dict[str, Any]] = {}
        try:
            for ws in wb.worksheets:
                headers = [c.value for c in ws[3] if c.value is not None] if ws.max_row >= 3 else []
                rows = []
                for r in range(4, ws.max_row + 1):
                    vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                    if any(v is not None for v in vals):
                        rows.append(vals)
                book[ws.title] = {"headers": headers, "rows": rows}
        finally:
            wb.close()
        return book

    def write_employee(self, code: str,
                       new_rows_by_sheet: Dict[str, List[Dict[str, Any]]]) -> Path:
        """เขียนไฟล์อัปเกรด (สำเนาต้นฉบับ + แถวใหม่) ลง output_dir.
        พยายามใช้ excel_io.write_employee (core-dev) ก่อน ถ้าไม่มี → fallback openpyxl.
        """
        src = Path(self.config["src_dir"]) / f"{code}_OneDrive_Profile.xlsx"
        dst = self.output_path(code)
        if dst.exists() and self.config.get("force"):
            dst.unlink()  # force = เริ่มจาก template ใหม่ ไม่ append ซ้ำแถวเดิม
        try:
            self._ensure_core_on_path()
            import excel_io  # ของ core-dev
            if hasattr(excel_io, "write_employee"):
                return excel_io.write_employee(
                    dst, append_rows_per_sheet=dict(new_rows_by_sheet), template_path=src)
        except Exception:
            pass
        return self._fallback_write(code, new_rows_by_sheet)

    def _fallback_write(self, code: str,
                        new_rows_by_sheet: Dict[str, List[Dict[str, Any]]]) -> Path:
        src = Path(self.config["src_dir"]) / f"{code}_OneDrive_Profile.xlsx"
        out_dir = Path(self.config["output_dir"])
        out_dir.mkdir(parents=True, exist_ok=True)
        dst = out_dir / f"{code}_OneDrive_Profile.xlsx"
        wb = openpyxl.load_workbook(src)
        try:
            for sheet, rows in new_rows_by_sheet.items():
                if sheet not in wb.sheetnames:
                    continue
                ws = wb[sheet]
                header_row = 3
                # สร้าง index: ชื่อคอลัมน์ -> เลขคอลัมน์ (1-based) แล้วขยาย header ด้วยคอลัมน์ร่วม
                col_idx: Dict[str, int] = {}
                for c in ws[header_row]:
                    if c.value is not None and str(c.value).strip():
                        col_idx[str(c.value).strip()] = c.column
                for cn in COMMON_COLS:
                    if cn not in col_idx:
                        col_idx[cn] = ws.max_column + 1
                        ws.cell(row=header_row, column=col_idx[cn], value=cn)
                for rec in rows:
                    rn = ws.max_row + 1
                    for name, val in rec.items():
                        ci = col_idx.get(name)
                        if ci:
                            ws.cell(row=rn, column=ci, value=val)
        finally:
            wb.save(dst)
            wb.close()
        return dst

    def output_path(self, code: str) -> Path:
        return Path(self.config["output_dir"]) / f"{code}_OneDrive_Profile.xlsx"

    # ------------------------------------------------------------------
    # Pair registry — รับประกันแถวครบ 2 ฝั่ง (eventId + logDateTime เดียวกัน)
    # ------------------------------------------------------------------
    def register_or_emit_pair(self, emp: Dict[str, Any], other_code: str, *,
                              event_id: str, log_dt: str, sheet: str, subject: str,
                              description: str, relationship: str, faction: str,
                              log_type: str, category: str, risk_level: str,
                              location: str, source: str) -> Optional[Dict[str, Any]]:
        """ลงทะเบียนคู่ (A,B) สำหรับ (eventId, sheet, logDateTime) แล้ว emit ฝั่ง emp.
        ถ้าฝั่งนี้ emit ไปแล้ว → คืน None (ไม่ซ้ำ).
        """
        code = emp["code"]
        if other_code not in self.identity_by_code:
            return None
        reg_key = f"{_pair_key(code, other_code)}|{event_id}|{sheet}|{log_dt}"
        em_key = f"{code}|{reg_key}"
        with self.lock:
            if em_key in self.pair_emitted:
                return None
            rec = self.pair_registry.get(reg_key)
            if rec is None:
                # routine pair (eventId='ROUTINE') เป็น low-key → ไม่บังคับ mirror ครบ 2 ฝั่ง
                is_routine = (event_id == "ROUTINE")
                rec = {
                    "a": sorted([code, other_code])[0],
                    "b": sorted([code, other_code])[1],
                    "eventId": event_id, "sheet": sheet, "logDateTime": log_dt,
                    "subject": subject, "descriptionTH": description,
                    "relationship": relationship, "faction": faction,
                    "logType": log_type, "category": category, "riskLevel": risk_level,
                    "location": location, "source": source,
                    "mirrorRequired": not is_routine,
                    "mirrorEmployeeCode": ("" if is_routine else other_code),
                }
                self.pair_registry[reg_key] = rec
            self.pair_emitted.add(em_key)
        return self._pair_row_for(rec, code)

    def absorb_pending_pairs(self, emp: Dict[str, Any],
                             plans: List[Dict[str, Any]]) -> int:
        """ดู pair_registry ว่ามีคู่ที่คนนี้เป็นฝั่ง B/C แล้วยังไม่ได้ emit → สร้างฝั่ง mirror
        (ใช้ timestamp/eventId เดียวกับฝั่งแรก) — เรียกก่อนวางแผนดราม่า."""
        code = emp["code"]
        added = 0
        with self.lock:
            for reg_key, rec in list(self.pair_registry.items()):
                if code not in (rec["a"], rec["b"]):
                    continue
                em_key = f"{code}|{reg_key}"
                if em_key in self.pair_emitted:
                    continue
                other = rec["b"] if rec["a"] == code else rec["a"]
                if other not in self.identity_by_code:
                    continue
                self.pair_emitted.add(em_key)
                plans.append(self._pair_row_for(rec, code))
                added += 1
        return added

    @staticmethod
    def _pair_row_for(rec: Dict[str, Any], code: str) -> Dict[str, Any]:
        other = rec["b"] if rec["a"] == code else rec["a"]
        return {
            "sheet": rec["sheet"],
            "kind": "drama_pair",
            "eventId": rec["eventId"],
            "logDateTime": rec["logDateTime"],
            "orig": {},
            "common": {
                "logDateTime": rec["logDateTime"],
                "logType": rec["logType"],
                "subject": rec["subject"],
                "counterpartyEmployeeCode": other,
                "eventId": rec["eventId"],
                "location": rec["location"],
                "source": rec["source"],
                "notes": rec["descriptionTH"],
            },
            "meta": {
                "employeeCode": code, "category": rec["category"],
                "riskLevel": rec["riskLevel"], "relationship": rec["relationship"],
                "faction": rec["faction"], "descriptionTH": rec["descriptionTH"],
                "logType": rec["logType"],
                "mirrorRequired": rec.get("mirrorRequired", False),
                "mirrorEmployeeCode": rec.get("mirrorEmployeeCode", ""),
            },
            "pair": {"other": other, "relationship": rec["relationship"],
                      "faction": rec["faction"]},
        }

    # ------------------------------------------------------------------
    # Drama pair registry — pair-centric (QA fix: mirror ครบ 2 ฝั่งตั้งแต่ต้น)
    #
    # หลักการ (ตาม QA root cause 1&2):
    #   - participant set ของแต่ละ event = canonical set เดียว
    #     (keyPairCodes + suggestedParticipants + คู่ใน matrix ที่อ้าง eventId)
    #     → ทุกคนที่เกี่ยวข้องใช้ set นี้ชุดเดียวกัน (ลบตัวเองออกตอน emit)
    #   - ทุกแถว drama ที่มี counterparty → ลงทะเบียน occurrence ใน drama_pending
    #     (eventId + sheet + logDateTime เดียวกัน) แล้ว emit ครบทุกฝั่ง:
    #       * คนที่ generate ทีหลัง  → absorb_drama_pending() บังคับสร้างฝั่ง mirror
    #       * คนที่ generate ทีก่อน  → backfill_drama_pending() (หลังวางแผนครบทุกคน)
    #   - แถวที่ถูกคนอื่นอ้างถึง = mandatory → ไม่โดน budget cap ตัด
    # ------------------------------------------------------------------

    def _file_lock(self, code: str) -> threading.Lock:
        """lock ต่อคน — serialize การเขียนไฟล์คนเดียวกัน (generate + backfill, workers>1)."""
        with self.lock:
            lk = self.file_locks.get(code)
            if lk is None:
                lk = threading.Lock()
                self.file_locks[code] = lk
            return lk

    def drama_participant_set(self, ev: Dict[str, Any], t: str) -> List[str]:
        """participant set canonical ของ event ณ timestamp t
        = keyPairCodes + suggestedParticipants + คู่ใน matrix ที่อ้าง eventId
        (กรองคนที่ยังไม่เข้าทำงาน ณ เวลานั้น — legacy 1997/2011 ไม่กรอง)
        คำนวณครั้งเดียวต่อ (eventId, t) แล้ว cache — ทุกคนใช้ชุดเดียวกัน."""
        key = f"{ev['eventId']}|{t}"
        with self.lock:
            cached = self.drama_participant_cache.get(key)
            if cached is not None:
                return cached
        codes: Set[str] = set()
        for a, b in ev.get("keyPairCodes", []):
            codes.add(a)
            codes.add(b)
        codes.update(ev.get("suggestedParticipants", []))
        for p in self.matrix.get("pairs", []):
            if ev["eventId"] in p.get("eventIds", []):
                codes.add(p["a"])
                codes.add(p["b"])
        is_legacy = ev.get("period") in ("1997", "2011")
        out: List[str] = []
        for c in sorted(codes):
            if c not in self.identity_by_code:
                continue
            hire = self.hire_date(c)
            if hire and not is_legacy and t[:10] < hire.isoformat():
                continue
            out.append(c)
        with self.lock:
            self.drama_participant_cache[key] = out
        return out

    def register_drama_occurrence(self, emp: Dict[str, Any], ev: Dict[str, Any], sheet: str,
                                  t: str, *, subject: str, notes: str, log_type: str,
                                  category: str, risk_level: str, location: str, source: str,
                                  rows_per: int = 1) -> Dict[str, Any]:
        """ลงทะเบียน occurrence ของ drama event (canonical) — mark ฝั่ง emp ว่า emit แล้ว.

        คืน record ของ occurrence; participants ใน record = canonical set เดียว
        ที่ทุกฝั่งต้อง emit (คนอื่นสร้าง mirror จาก record นี้ด้วย eventId+logDateTime เดียวกัน).
        """
        code = emp["code"]
        occ_key = f"{ev['eventId']}|{sheet}|{t}"
        participants = self.drama_participant_set(ev, t)
        with self.lock:
            rec = self.drama_pending.get(occ_key)
            if rec is None:
                rec = {
                    "eventId": ev["eventId"], "sheet": sheet, "logDateTime": t,
                    "participants": participants,
                    "subject": subject, "notes": notes,
                    "logType": log_type, "category": category,
                    "riskLevel": risk_level, "location": location, "source": source,
                    "rowsPer": rows_per,
                }
                self.drama_pending[occ_key] = rec
            self.drama_side_emitted.add(f"{code}|{occ_key}")
        return rec

    def _drama_row_from_record(self, rec: Dict[str, Any], code: str) -> List[Dict[str, Any]]:
        """สร้างแถว drama (ฝั่ง mirror) จาก occurrence record — eventId/logDateTime
        ใช้ค่าเดียวกับฝั่งต้นทาง; counterparties = participants - {code} (canonical set เดียว)."""
        ev = self.event_index.get(rec["eventId"], {})
        emp = self.identity_by_code.get(code, {})
        participants = [p for p in rec.get("participants", []) if p != code]
        cp_str = ";".join(participants[:6])
        rows: List[Dict[str, Any]] = []
        for _ in range(int(rec.get("rowsPer", 1))):
            rows.append({
                "sheet": rec["sheet"], "kind": "drama_pair", "eventId": rec["eventId"],
                "logDateTime": rec["logDateTime"],
                "orig": _drama_orig_values(rec["sheet"], emp, ev, self, rec["logDateTime"], participants),
                "common": {
                    "logDateTime": rec["logDateTime"], "logType": rec["logType"],
                    "subject": rec["subject"], "counterpartyEmployeeCode": cp_str,
                    "eventId": rec["eventId"], "location": rec["location"],
                    "source": rec["source"], "notes": rec["notes"],
                },
                "meta": {"employeeCode": code, "category": rec["category"],
                         "riskLevel": rec["riskLevel"], "relationship": "",
                         "faction": self.faction_by_code.get(code, ""),
                         "descriptionTH": rec["notes"], "logType": rec["logType"],
                         "mirrorRequired": True, "mirrorEmployeeCode": cp_str},
                "pair": None,
            })
        return rows

    def absorb_drama_pending(self, emp: Dict[str, Any],
                             plans: List[Dict[str, Any]]) -> int:
        """สร้างแถว drama ที่เป็น mandatory สำหรับ emp — occurrence ที่คนอื่นลงทะเบียนไว้
        แล้ว emp อยู่ใน participant set (ถูกอ้างถึง) → บังคับสร้างฝั่ง mirror
        (ไม่ถูก budget cap ตัด — เรียกก่อน _plan_drama_rows)."""
        code = emp["code"]
        added = 0
        with self.lock:
            occ_items = list(self.drama_pending.items())
        for occ_key, rec in occ_items:
            if code not in rec.get("participants", []):
                continue
            side_key = f"{code}|{occ_key}"
            with self.lock:
                if side_key in self.drama_side_emitted:
                    continue
                self.drama_side_emitted.add(side_key)
            plans.extend(self._drama_row_from_record(rec, code))
            added += 1
        return added

    def _inject_record_for_plan(self, code: str, plan: Dict[str, Any]) -> None:
        """บันทึก metadata (injected_events) ของแถวที่มี eventId/คู่คน — Phase 3 ใช้ตรวจข้ามไฟล์."""
        if not (plan.get("eventId") or plan.get("pair") or plan.get("pairWith")):
            return
        common = plan["common"]
        rec = {
            "eventId": plan.get("eventId", ""),
            "logDateTime": plan["logDateTime"],
            "sheet": plan["sheet"],
            "employeeCode": code,
            "counterpartyEmployeeCode": common.get("counterpartyEmployeeCode", ""),
            "subject": common.get("subject", ""),
            "descriptionTH": common.get("notes", ""),
            "riskLevel": (plan.get("meta") or {}).get("riskLevel", ""),
            "category": (plan.get("meta") or {}).get("category", ""),
            "logType": common.get("logType", ""),
            "source": common.get("source", ""),
            "location": common.get("location", ""),
            "relationship": (plan.get("meta") or {}).get("relationship", ""),
            "faction": (plan.get("meta") or {}).get("faction", ""),
            "rowKind": ("routine_pair" if plan.get("eventId") == "ROUTINE"
                        else plan.get("kind", "drama")),
            "mirrorRequired": (plan.get("meta") or {}).get("mirrorRequired", False),
            "mirrorEmployeeCode": (plan.get("meta") or {}).get("mirrorEmployeeCode", ""),
        }
        self.append_injected(rec)

    def _write_rows_locked(self, code: str, rows_by_sheet: Dict[str, List[Dict[str, Any]]]) -> int:
        """เขียนแถวลงไฟล์ของ code ภายใต้ per-file lock (กันชนกับ generate/backfill คนอื่น)."""
        if not rows_by_sheet:
            return 0
        n = sum(len(v) for v in rows_by_sheet.values())
        with self._file_lock(code):
            try:
                self.write_employee(code, dict(rows_by_sheet))
            except Exception:
                return 0
        return n

    def backfill_drama_pending(self) -> int:
        """หลังวางแผนครบทุกคน: เขียนฝั่ง mirror ที่ยังขาด (ถูกคนอื่นอ้างถึงแต่ยังไม่มีแถว —
        เช่น เกิดก่อนคนที่อ้าง / วิ่ง parallel) ลงไฟล์ Excel ของฝั่งนั้นโดยตรง.
        ใช้ (eventId, logDateTime, participants) เดียวกับ occurrence — idempotent (side_key)."""
        added = 0
        with self.lock:
            occ_items = list(self.drama_pending.items())
        for occ_key, rec in occ_items:
            for code in rec.get("participants", []):
                side_key = f"{code}|{occ_key}"
                with self.lock:
                    if side_key in self.drama_side_emitted:
                        continue
                # คนที่กำลัง generate อยู่ → ข้าม (คนนั้นจะ emit เองในรอบของตัวเอง)
                if code in self._generating:
                    continue
                # เขียนเฉพาะคนที่ถูก generate แล้ว (ไฟล์ output มีอยู่จริง) —
                # ห้ามสร้างไฟล์ใหม่ให้คนนอก scope (กัน backfill ไปสร้างไฟล์ทิ้งทั้งชุด)
                if not self.output_path(code).exists():
                    continue
                # mark ฝั่งว่า emit หลังผ่าน checks ทั้งหมดแล้ว (กัน thread อื่นเขียนซ้ำ)
                with self.lock:
                    if side_key in self.drama_side_emitted:
                        continue
                    self.drama_side_emitted.add(side_key)
                rows = self._drama_row_from_record(rec, code)
                if not rows:
                    continue
                by_sheet: Dict[str, List[Dict[str, Any]]] = {rec["sheet"]: rows}
                n = self._write_rows_locked(code, by_sheet)
                if n <= 0:
                    continue
                for r in rows:
                    self._inject_record_for_plan(code, r)
                added += n
        return added

    # ------------------------------------------------------------------
    # Progress / Checkpoint (idempotent ข้าม run)
    # ------------------------------------------------------------------
    def log_progress(self, entry: Dict[str, Any]) -> None:
        with self.lock:
            self.progress_log.append(entry)
            self.stats["employees_done"] += 1
            path = Path(self.config["output_dir"]).parent / "progress.jsonl"
            path.parent.mkdir(parents=True, exist_ok=True)
            with open(path, "a", encoding="utf-8") as f:
                f.write(json.dumps(entry, ensure_ascii=False) + "\n")

    def _checkpoint_dir(self) -> Path:
        d = Path(self.config["output_dir"]).parent
        d.mkdir(parents=True, exist_ok=True)
        return d

    def _load_checkpoint(self) -> None:
        d = self._checkpoint_dir()
        try:
            ev = json.loads((d / "checkpoint_events.json").read_text("utf-8"))
            self.event_times = ev.get("event_times", {})
            self.drama_descriptions = {k: v for k, v in ev.get("drama_descriptions", {}).items()}
            self.drama_emitted = set(ev.get("drama_emitted", []))
            self.pair_emitted = set(ev.get("pair_emitted", []))
            self.pair_registry = ev.get("pair_registry", {})
            self.drama_pending = ev.get("drama_pending", {})
            self.drama_side_emitted = set(ev.get("drama_side_emitted", []))
        except FileNotFoundError:
            pass
        try:
            inj = (d / "injected_events.jsonl")
            if inj.exists():
                self.injected_events = [json.loads(line) for line in inj.read_text("utf-8").splitlines() if line.strip()]
        except FileNotFoundError:
            pass

    def save_checkpoint(self) -> None:
        d = self._checkpoint_dir()
        # snapshot ใต้ lock — ป้องกัน dict เปลี่ยนขนาดระหว่าง json.dump เมื่อ workers > 1
        with self.lock:
            payload = {
                "event_times": dict(self.event_times),
                "drama_descriptions": {k: dict(v) for k, v in self.drama_descriptions.items()},
                "drama_emitted": sorted(self.drama_emitted),
                "pair_emitted": sorted(self.pair_emitted),
                "pair_registry": dict(self.pair_registry),
                "drama_pending": dict(self.drama_pending),
                "drama_side_emitted": sorted(self.drama_side_emitted),
            }
        with open(d / "checkpoint_events.json", "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=1)

    def append_injected(self, record: Dict[str, Any]) -> None:
        with self.lock:
            self.injected_events.append(record)
            path = self._checkpoint_dir() / "injected_events.jsonl"
            with open(path, "a", encoding="utf-8") as f:
                f.write(json.dumps(record, ensure_ascii=False) + "\n")

# ---------------------------------------------------------------------------
# Planning helpers
# ---------------------------------------------------------------------------
def _drama_orig_values(sheet: str, emp: Dict[str, Any], ev: Dict[str, Any], ctx: DataGenContext,
                       t: str, counterparties: List[str]) -> Dict[str, Any]:
    """ค่าเติมคอลัมน์เดิมของ sheet สำหรับแถวดราม่า (สมจริงตามประเภทเหตุการณ์)."""
    code = emp["code"]
    rng = ctx.get_rng()
    risk = ev.get("riskLevel", "medium")
    severity = {"critical": "Critical", "high": "High", "medium": "Moderate", "low": "Low"}.get(risk, "Moderate")
    cat = ev.get("category", "")
    dt = datetime.fromisoformat(t.replace("+07:00", ""))
    t1 = emp.get("department", "") in ("Executive", "HR & Admin", "Finance & Accounting", "Legal")
    if sheet == "Warning_Disciplinary_History":
        return {"caseId": f"WARN-{ev['eventId']}-{code}", "caseDate": t[:10],
                "caseType": (ev.get("titleTH", "") or "")[:45], "severity": severity,
                "formalWarning": "Yes" if risk in ("high", "critical") else "No",
                "summary": ev.get("titleTH", ""), "rootCause": "อยู่ระหว่างการสอบข้อเท็จจริง",
                "actionTaken": "ตักเตือน/เรียกชี้แจง", "resolutionStatus": ev.get("resolutionStatus", "Open"),
                "managerInvolved": emp.get("managerName", ""),
                "hrConfidentialityLevel": "Tier 1 - Strict" if t1 else "Tier 2 - Sensitive",
                "redactionRequired": "Yes" if t1 else "No",
                "linkedProjectId": "", "linkedTrainingId": ""}
    if sheet == "Grievance_Log":
        return {"Complaint_Type": (ev.get("titleTH", "") or "")[:40],
                "Status": ev.get("resolutionStatus", "Open")}
    if sheet == "Expense_Reports":
        base = 2000 if cat == "grey_area_collusion" else 300
        return {"Travel_THB": rng.randint(base, base + 8000),
                "Entertainment_THB": rng.randint(base, base + 6000),
                "Office_Supplies_THB": rng.randint(base, base + 4000)}
    if sheet == "KPI_OKR_History":
        neg = cat in ("dept_negative", "grey_area_collusion", "cross_dept_conflict")
        score = round(rng.uniform(1.5, 2.8), 1) if neg else round(rng.uniform(3.8, 4.6), 1)
        band = "Unsatisfactory (E)" if score < 2 else ("Below (D)" if score < 2.8 else "Meets (C)")
        return {"reviewPeriod": f"{dt.year}-Q{(dt.month - 1) // 3 + 1}", "kpiScore": score,
                "okrScore": round(max(1.0, score - 0.4), 1), "performanceBand": band,
                "strongArea": "", "weakArea": "Performance Consistency" if neg else "",
                "managerFeedback": ("ต้องปรับปรุงอย่างเร่งด่วน — มีการกำหนด PIP และติดตามอย่างใกล้ชิด"
                                    if band in ("Unsatisfactory (E)", "Below (D)") else "ผลงานเป็นไปตามเป้าหมาย"),
                "improvementPlan": ("กำหนด Performance Improvement Plan 60 วัน"
                                    if band in ("Unsatisfactory (E)", "Below (D)") else ""),
                "followUpStatus": "In Progress" if band in ("Unsatisfactory (E)", "Below (D)") else ""}
    if sheet == "Project_History":
        neg = cat in ("dept_negative", "grey_area_collusion", "cross_dept_conflict", "crisis")
        return {"projectId": rng.choice(PROJECT_IDS), "role": "Project Lead",
                "contributionSummary": ev.get("titleTH", ""),
                "individualOutcome": "ล่าช้า/มีปัญหา" if neg else "สำเร็จตามแผน",
                "hasMistake": "Yes" if neg else "No",
                "mistakeIssue": ev.get("titleTH", ""),
                "recoveryAction": "แผนแก้ไขอยู่ระหว่างดำเนินการ" if neg else ""}
    if sheet == "Career_Timeline":
        return {"date": t[:10], "eventType": "Review", "title": ev.get("titleTH", ""),
                "department": emp.get("department", ""),
                "notes": (ev.get("descriptionTH", "") or "")[:120]}
    if sheet == "360_Feedback":
        return {"Reviewer_Type": "Manager", "Comment": ev.get("titleTH", "")}
    if sheet == "Timesheet_Log":
        b = rng.randint(40, 90)
        return {"Billable_Hours_Pct": b, "Admin_Hours_Pct": 100 - b}
    if sheet == "Attendance_Record":
        return {"Sick_Leave_Days": 0, "Personal_Leave_Days": 0, "Late_Arrivals": rng.randint(0, 3)}
    if sheet == "IT_Ticket_Log":
        # sync Status กับ resolutionStatus ใน catalog (QA bonus): Ongoing → Open ไม่ใช่ Resolved
        rs = (ev.get("resolutionStatus", "") or "").lower()
        if "ongoing" in rs or "in progress" in rs:
            status = "Open"
        elif "investigat" in rs:
            status = "Under Investigation"
        elif "closed" in rs or "resolved" in rs:
            status = "Resolved"
        elif rs:
            status = ev["resolutionStatus"]
        else:
            status = "Under Investigation" if risk in ("high", "critical") else "Resolved"
        return {"Ticket_Issue": (ev.get("titleTH", "") or "")[:40],
                "Status": status}
    if sheet == "Compliance_Mandates":
        return {"Mandate": (ev.get("titleTH", "") or "")[:40],
                "Status": "Under Review" if risk in ("high", "critical") else "Compliant"}
    # fallback: ค่า routine ทั่วไป
    return orig_values_for(sheet, emp, ctx, rng, dt, {"activity": ev.get("titleTH", "")})


def _collab_orig(emp: Dict[str, Any], ctx: DataGenContext, other: str, rel: str) -> Dict[str, Any]:
    other_e = ctx.identity_by_code.get(other, {})
    return {"collaboratorEmployeeId": other_e.get("pk", ""), "collaboratorName": other_e.get("name", ""),
            "collaboratorDept": other_e.get("department", ""), "projectId": ctx.get_rng().choice(PROJECT_IDS),
            "relationshipType": REL_NAME_TH.get(rel, rel),
            "collaborationQuality": "Difficult" if rel in ("conflict", "collusion") else ctx.get_rng().choice(["Good", "Good", "Normal"]),
            "hasConflict": "Yes" if rel in ("conflict", "collusion") else "No",
            "conflictSummary": "ความขัดแย้งที่เกี่ยวข้องกับเหตุการณ์" if rel in ("conflict", "collusion") else "",
            "resolutionSummary": "อยู่ระหว่างการไกล่เกลี่ย" if rel in ("conflict", "collusion") else ""}


def _build_row_values(orig_headers: List[str], orig: Dict[str, Any], common: Dict[str, Any]) -> Dict[str, Any]:
    """เรียงค่าของแถวใหม่ให้ตรงกับ full header (คอลัมน์เดิม + คอลัมน์ร่วม)
    คืน dict {header_name: value} — รองรับทั้ง excel_io ของ core-dev (append_rows)
    และ fallback writer ของเรา"""
    full_headers = list(orig_headers) + [c for c in COMMON_COLS if c not in orig_headers]
    out: Dict[str, Any] = {}
    for h in full_headers:
        if h in COMMON_COLS:
            out[h] = common.get(h, "")
        else:
            out[h] = orig.get(h, "")
    return out

# ---------------------------------------------------------------------------
# Plan builders
# ---------------------------------------------------------------------------
def _plan_drama_rows(emp: Dict[str, Any], ctx: DataGenContext,
                     cap: Optional[int] = None,
                     hard_cap: Optional[int] = None) -> List[Dict[str, Any]]:
    """แถวดราม่า (20%) — pair-centric (QA fix: mirror ครบ 2 ฝั่งตั้งแต่ต้น)

    - participant set ของ event = canonical set เดียว (keyPair + suggested + คู่ใน matrix
      ที่อ้าง eventId) → ทุกคนที่เกี่ยวข้องใช้ชุดเดียวกัน (ลบตัวเองออกตอน emit)
    - ทุกแถว drama ที่มี counterparty → ลงทะเบียน occurrence ใน ``ctx.drama_pending``
      (eventId + sheet + logDateTime เดียวกัน) → ฝั่งอื่นถูกบังคับให้ mirror
      (absorb_drama_pending / backfill_drama_pending)
    - เหตุการณ์ที่ emp เป็นตัวละครหลัก (keyPair/critical) หรือถูกคนอื่นอ้างถึง = mandatory
      → ไม่โดน budget cap ตัด (แก้ root cause 2: POL-05 ถูก drop ฝั่ง primary)
    - ``cap``/``hard_cap`` ควบคุมเฉพาะ event ที่คนนี้เป็นคน "เริ่ม" (initiated) เท่านั้น

    (แถว Collaboration_Network แยกไปใน _plan_collab_pairs เพื่อให้ครบ 2 ฝั่งเสมอ)
    """
    code = emp["code"]
    hire = ctx.hire_date(code)
    plans: List[Dict[str, Any]] = []
    if cap is None:
        cap = ctx.drama_cap(emp)
    # hard ceiling ใช้กับ event ที่ emp เป็นคนเริ่มเท่านั้น — mandatory ห้ามตัด
    if hard_cap is None:
        hard_cap = max(60, int(round(ctx.drama_cap(emp) * 1.25)))

    def _mandatory(ev: Dict[str, Any]) -> bool:
        """เหตุการณ์ที่ emp เป็น 'ตัวละครหลัก' (keyPairCodes) หรือ crisis+critical+suggested
        → ต้องได้แถวครบ ไม่ถูกตัดด้วยงบ (สอดคล้อง DESIGN: keyPairCodes บังคับ)"""
        if any(code in (a, b) for a, b in ev.get("keyPairCodes", [])):
            return True
        if ev.get("riskLevel") == "critical" and code in ev.get("suggestedParticipants", []):
            return True
        return False

    initiated = 0  # จำนวน event ที่คนนี้เริ่ม (budgeted) — mandatory/absorbed ไม่นับ
    for ev in ctx.events_for(code):
        if ev.get("category") == "routine":
            continue  # WORK-* ไปจัดการใน _plan_catalog_routine_rows
        mandatory = _mandatory(ev)
        if not mandatory and initiated >= cap:
            continue  # งบเต็ม — ข้าม event ใหม่ แต่ยังเช็ค event ที่เหลือว่า mandatory ไหม
        eid = ev["eventId"]
        is_legacy = ev.get("period") in ("1997", "2011")
        times = ctx.ensure_event_times(ev)
        sheets = [SHEET_ALIAS.get(s, s) for s in ev.get("affectedSheets", [])]
        sheets = [s for s in sheets if s in VALID_SHEETS and s != "Collaboration_Network"]
        for t in times:
            if hire and not is_legacy and t[:10] < hire.isoformat():
                continue
            participants = ctx.drama_participant_set(ev, t)   # canonical set เดียว
            counterparties = [p for p in participants if p != code]
            cp_str = ";".join(counterparties[:6])
            rows_per = 2 if (int(ev.get("logRowExpansion", 1)) >= 4 or ev.get("riskLevel") in ("high", "critical")) else 1
            for sheet in sheets:
                occ_key = f"{eid}|{sheet}|{t}"
                side_key = f"{code}|{occ_key}"
                with ctx.lock:
                    if side_key in ctx.drama_side_emitted:
                        continue  # absorb สร้างไปแล้ว (ถูกคนอื่นอ้างถึง) — ไม่ซ้ำ
                # เหตุการณ์ที่ถูกคนอื่นอ้างถึง (occurrence ลงทะเบียนแล้ว) → mandatory เช่นกัน
                is_mandatory = mandatory or occ_key in ctx.drama_pending
                if not is_mandatory and initiated >= cap:
                    continue
                if not is_mandatory and initiated >= hard_cap:
                    continue
                sdesc = ctx.ensure_drama_description(ev, sheet, {
                    "employeeCode": code, "employeeName": emp["name"],
                    "counterpartyEmployeeCode": cp_str,
                    "counterpartyNames": ";".join(ctx.identity_by_code.get(c, {}).get("name", c) for c in counterparties[:6])})
                log_type = LOG_TYPE_BY_CATEGORY.get(ev.get("category", ""), "incident")
                # ลงทะเบียน occurrence (canonical) → ทุกฝั่งใน participants ต้อง mirror
                ctx.register_drama_occurrence(
                    emp, ev, sheet, t, subject=sdesc["subject"], notes=sdesc["descriptionTH"],
                    log_type=log_type, category=ev.get("category", ""),
                    risk_level=ev.get("riskLevel", "medium"),
                    location=ctx._event_location(ev),
                    source=SHEET_SOURCE.get(sheet, "HRIS"), rows_per=rows_per)
                for _k in range(rows_per):
                    plans.append({
                        "sheet": sheet, "kind": "drama_pair", "eventId": eid, "logDateTime": t,
                        "orig": _drama_orig_values(sheet, emp, ev, ctx, t, counterparties),
                        "common": {
                            "logDateTime": t, "logType": log_type,
                            "subject": sdesc["subject"],
                            "counterpartyEmployeeCode": cp_str, "eventId": eid,
                            "location": ctx._event_location(ev),
                            "source": SHEET_SOURCE.get(sheet, "HRIS"),
                            "notes": sdesc["descriptionTH"],
                        },
                        "meta": {"employeeCode": code, "category": ev.get("category", ""),
                                 "riskLevel": ev.get("riskLevel", "medium"),
                                 "relationship": "", "faction": ctx.faction_by_code.get(code, ""),
                                 "descriptionTH": sdesc["descriptionTH"], "logType": log_type,
                                 "mirrorRequired": True, "mirrorEmployeeCode": cp_str},
                        "pair": None,
                    })
                if not is_mandatory:
                    initiated += 1
    return plans


def _plan_collab_pairs(emp: Dict[str, Any], ctx: DataGenContext) -> List[Dict[str, Any]]:
    """แถว Collaboration_Network ต่อคู่ — DESIGN.md §9: ทุก pair มีแถวครบ 2 ฝั่ง
    ด้วย eventId + logDateTime เดียวกัน (timestamp มาจาก ctx.event_times ที่ global
    → deterministic ทั้ง 2 ฝั่ง ไม่ขึ้นกับลำดับการ generate).

    ต่อคู่: ใช้ eventId แรกของ pair (1 occurrence แรก) — ควบคุมปริมาณให้สมมาตร
    """
    code = emp["code"]
    hire = ctx.hire_date(code)
    plans: List[Dict[str, Any]] = []
    for pair in ctx.pairs_by_code.get(code, []):
        other = pair["b"] if pair["a"] == code else pair["a"]
        if other not in ctx.identity_by_code:
            continue
        eids = [e for e in pair.get("eventIds", []) if e in ctx.event_index]
        if not eids:
            continue
        ev = ctx.event_index[eids[0]]
        eid = ev["eventId"]
        is_legacy = ev.get("period") in ("1997", "2011")
        times = ctx.ensure_event_times(ev)
        t = times[0]
        if hire and not is_legacy and t[:10] < hire.isoformat():
            continue
        other_hire = ctx.hire_date(other)
        if other_hire and not is_legacy and t[:10] < other_hire.isoformat():
            continue
        rel = pair.get("relationship", "work_partner")
        em_key = f"{code}|{_pair_key(code, other)}|{eid}|Collaboration_Network|{t}"
        with ctx.lock:
            if em_key in ctx.pair_emitted:
                continue
            ctx.pair_emitted.add(em_key)
        cdesc = ctx.ensure_drama_description(ev, "Collaboration_Network", {
            "employeeCode": code, "employeeName": emp["name"],
            "counterpartyEmployeeCode": other,
            "counterpartyNames": ctx.identity_by_code.get(other, {}).get("name", other)})
        plans.append({
            "sheet": "Collaboration_Network", "kind": "drama_pair", "eventId": eid,
            "logDateTime": t,
            "orig": _collab_orig(emp, ctx, other, rel),
            "common": {
                "logDateTime": t,
                "logType": LOG_TYPE_BY_CATEGORY.get(ev.get("category", ""), "incident"),
                "subject": f"{cdesc['subject']} — {REL_NAME_TH.get(rel, rel)}",
                "counterpartyEmployeeCode": other, "eventId": eid,
                "location": ctx._event_location(ev), "source": "OrgGraph",
                "notes": cdesc["descriptionTH"],
            },
            "meta": {"employeeCode": code, "category": ev.get("category", ""),
                     "riskLevel": ev.get("riskLevel", "medium"), "relationship": rel,
                     "faction": pair.get("faction", ""), "descriptionTH": cdesc["descriptionTH"],
                     "logType": LOG_TYPE_BY_CATEGORY.get(ev.get("category", ""), "incident"),
                     "mirrorRequired": True, "mirrorEmployeeCode": other},
            "pair": {"other": other, "relationship": rel, "faction": pair.get("faction", "")},
        })
    return plans


def _plan_catalog_routine_rows(emp: Dict[str, Any], ctx: DataGenContext) -> List[Dict[str, Any]]:
    """เหตุการณ์ routine จาก catalog (WORK-01..08, recurring) — มี eventId แต่ logType=routine."""
    code = emp["code"]
    hire = ctx.hire_date(code)
    plans: List[Dict[str, Any]] = []
    for ev in ctx.events_for(code):
        if ev.get("category") != "routine":
            continue
        if code not in ev.get("suggestedParticipants", []):
            continue
        eid = ev["eventId"]
        sheets = [SHEET_ALIAS.get(s, s) for s in ev.get("affectedSheets", [])]
        sheets = [s for s in sheets if s in VALID_SHEETS and s != "Collaboration_Network"]
        for t in ctx.ensure_event_times(ev):
            if hire and t[:10] < hire.isoformat():
                continue
            for sheet in sheets:
                dkey = f"{eid}|{code}|{sheet}|{t}"
                with ctx.lock:
                    if dkey in ctx.drama_emitted:
                        continue
                    ctx.drama_emitted.add(dkey)
                sdesc = ctx.ensure_drama_description(ev, sheet, {
                    "employeeCode": code, "employeeName": emp["name"]})
                plans.append({
                    "sheet": sheet, "kind": "catalog_routine", "eventId": eid, "logDateTime": t,
                    "orig": _drama_orig_values(sheet, emp, ev, ctx, t, []),
                    "common": {
                        "logDateTime": t, "logType": "routine",
                        "subject": sdesc["subject"], "counterpartyEmployeeCode": "", "eventId": eid,
                        "location": ctx._event_location(ev),
                        "source": SHEET_SOURCE.get(sheet, "HRIS"),
                        "notes": sdesc["descriptionTH"],
                    },
                    "meta": {"employeeCode": code, "category": "routine",
                             "riskLevel": ev.get("riskLevel", "low"), "relationship": "",
                             "faction": ctx.faction_by_code.get(code, ""),
                             "descriptionTH": sdesc["descriptionTH"], "logType": "routine",
                             "mirrorRequired": False, "mirrorEmployeeCode": ""},
                    "pair": None,
                })
    return plans

# ---------------------------------------------------------------------------
# generate_employee — หัวใจ: 1 คน ต่อ 1 ไฟล์
# ---------------------------------------------------------------------------
def generate_employee(emp: Dict[str, Any], ctx: DataGenContext) -> Dict[str, Any]:
    """สร้าง rows สำหรับพนักงาน 1 คน → เขียน Excel ลง output_dir.

    Return dict: {employeeCode, rowsAdded, sheetsUpdated, injectedEvents,
                  dramaRows, routineRows, finalRows, skipped?}
    """
    code = emp["code"]
    if ctx.output_path(code).exists() and not ctx.config.get("force"):
        return {"employeeCode": code, "skipped": True, "reason": "already_generated",
                "rowsAdded": 0, "sheetsUpdated": [], "injectedEvents": 0,
                "dramaRows": 0, "routineRows": 0, "finalRows": 0}

    # กัน thread อื่น backfill เขียนไฟล์คนนี้ตอนกำลัง generate (workers > 1)
    ctx._generating.add(code)

    book = ctx.read_employee(code)
    existing_total = sum(len(s["rows"]) for s in book.values())

    # Batch-fetch คำอธิบาย (eventId, sheet) ที่คนนี้ต้องใช้ → 1 API call ต่อ N events
    # (offline: fallback template ต่อ spec — ผลเหมือนเดิม; ทำก่อน planning ให้ cache เต็ม)
    ctx.prewarm_descriptions(emp)

    plans: List[Dict[str, Any]] = []
    absorbed = ctx.absorb_pending_pairs(emp, plans)          # routine mirror จากคู่ที่ generate ก่อน
    absorbed += ctx.absorb_drama_pending(emp, plans)         # drama mirror (mandatory — ไม่โดน cap ตัด)
    collab_plans = _plan_collab_pairs(emp, ctx)              # Collaboration_Network ครบ 2 ฝั่ง
    total_drama = ctx.drama_cap(emp)
    sheet_cap = max(20, total_drama - len(collab_plans))
    drama_plans = _plan_drama_rows(emp, ctx, cap=sheet_cap,
                                   hard_cap=max(60, int(round(total_drama * 1.25))))
    work_plans = _plan_catalog_routine_rows(emp, ctx)        # WORK-* จาก catalog
    plans += collab_plans + drama_plans + work_plans

    target = int(ctx.config["target_rows"])
    minr = int(ctx.config["min_rows"])
    maxr = int(ctx.config["max_rows"])
    new_so_far = len(plans)
    routine_budget = max(0, min(target, maxr) - existing_total - new_so_far)
    if existing_total + new_so_far < minr:
        routine_budget = max(routine_budget, minr - existing_total - new_so_far)
    routine_plans = generate_routine_plan(emp, ctx, routine_budget)   # routine 80%
    plans += routine_plans

    # safety: total ต้องไม่เกิน max_rows (ตัด routine ทิ้งก่อน เพราะต่อท้ายอยู่)
    max_new = maxr - existing_total
    if max_new < 0:
        plans = []
    elif len(plans) > max_new:
        plans = plans[:max_new]

    # ---- apply ลง workbook ----
    new_rows_by_sheet: Dict[str, List[Dict[str, Any]]] = collections.defaultdict(list)
    sheets_updated: Set[str] = set()
    injected_count = 0
    for plan in plans:
        sheet = plan["sheet"]
        if sheet not in book:
            continue
        orig_headers = book[sheet]["headers"] or SHEET_ORIGINAL_COLS.get(sheet, [])
        common = plan["common"]
        orig = dict(plan.get("orig") or {})
        if plan.get("kind") == "drama_pair" and not orig:  # ฝั่งคู่ → เติมคอลัมน์เดิม collab
            if sheet == "Collaboration_Network":
                other = common.get("counterpartyEmployeeCode", "")
                rel = (plan.get("pair") or {}).get("relationship", "work_partner")
                orig = _collab_orig(emp, ctx, other, rel)
        full = _build_row_values(orig_headers, orig, common)
        new_rows_by_sheet[sheet].append(full)
        sheets_updated.add(sheet)

        # routine pair rows → ลงทะเบียนคู่ (ให้อีกฝั่งสร้าง mirror เมื่อ generate ถึง)
        if plan.get("kind") == "routine" and plan.get("pairWith"):
            other = plan["pairWith"]
            rel, faction = "work_partner", ""
            for p in ctx.pairs_by_code.get(code, []):
                if other in (p["a"], p["b"]):
                    rel = p.get("relationship", "work_partner")
                    faction = p.get("faction", "")
                    break
            ctx.register_or_emit_pair(emp, other, event_id="ROUTINE", log_dt=plan["logDateTime"],
                                      sheet=sheet, subject=common.get("subject", ""),
                                      description=common.get("notes", ""), relationship=rel,
                                      faction=faction, log_type="routine", category="routine",
                                      risk_level="low", location=common.get("location", ""),
                                      source=common.get("source", ""))

        # injected_events — เฉพาะแถวที่มี eventId หรือคู่คน (Phase 3 ตรวจข้ามไฟล์)
        ctx._inject_record_for_plan(code, plan)
        if plan.get("eventId") or plan.get("pair") or plan.get("pairWith"):
            injected_count += 1

    # เขียนไฟล์ภายใต้ per-file lock (กันชนกับ backfill ของ thread อื่นเมื่อ workers > 1)
    ctx._write_rows_locked(code, dict(new_rows_by_sheet))
    # backfill: เขียนฝั่ง mirror ที่ถูกคนอื่นอ้างถึงแต่ยังไม่มีแถว (คนที่ generate ก่อนหน้าเรา)
    backfilled = ctx.backfill_drama_pending()
    ctx._generating.discard(code)
    final_total = existing_total + len(plans)
    drama_rows = len(collab_plans) + len(drama_plans) + len(work_plans) + absorbed
    ctx.log_progress({
        "employeeCode": code, "name": emp["name"], "department": emp.get("department", ""),
        "existingRows": existing_total, "rowsAdded": len(plans),
        "dramaRows": drama_rows, "routineRows": len(routine_plans),
        "sheetsUpdated": sorted(sheets_updated), "finalRows": final_total, "status": "ok",
    })
    ctx.save_checkpoint()
    return {
        "employeeCode": code, "rowsAdded": len(plans), "sheetsUpdated": sorted(sheets_updated),
        "injectedEvents": injected_count, "dramaRows": drama_rows,
        "routineRows": len(routine_plans), "finalRows": final_total,
        "backfilledMirrorRows": backfilled,
    }

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main(argv: Optional[List[str]] = None) -> List[Dict[str, Any]]:
    ap = argparse.ArgumentParser(description="Phase 2 Data Generation Engine — BuildersEye 360")
    ap.add_argument("--codes", default="", help="EMP001,EMP005,... (คั่น comma)")
    ap.add_argument("--all", action="store_true", help="generate ทุกคนใน identity-graph")
    ap.add_argument("--no-api", action="store_true", default=None,
                    help="ใช้ template offline (ไม่ยิง DeepSeek)")
    ap.add_argument("--api-key", default="", help="DeepSeek API key (หรือ env DEEPSEEK_API_KEY)")
    ap.add_argument("--workers", type=int, default=1, help="parallel threads (ค่าเริ่มต้น 1)")
    ap.add_argument("--target-rows", type=int, default=420)
    ap.add_argument("--min-rows", type=int, default=300)
    ap.add_argument("--max-rows", type=int, default=500)
    ap.add_argument("--desc-batch-size", type=int, default=8,
                    help="จำนวน events ต่อ 1 API call (batch ลด token overhead; offline ไม่กระทบ)")
    ap.add_argument("--seed", type=int, default=20260807)
    ap.add_argument("--force", action="store_true", help="ทับไฟล์ output เดิม (ถ้ามี)")
    ap.add_argument("--out-dir", default="", help="output dir (ค่าเริ่มต้น tools/upgrade_360/output/hr_onedrive_upgraded)")
    args = ap.parse_args(argv)

    env_key = os.environ.get("DEEPSEEK_API_KEY", "")
    no_api = args.no_api if args.no_api is not None else (not bool(args.api_key or env_key))
    config: Dict[str, Any] = {
        "no_api": no_api, "api_key": args.api_key or env_key, "workers": args.workers,
        "target_rows": args.target_rows, "min_rows": args.min_rows, "max_rows": args.max_rows,
        "seed": args.seed, "force": args.force, "desc_batch_size": args.desc_batch_size,
    }
    if args.out_dir:
        config["output_dir"] = args.out_dir

    ctx = DataGenContext(config)
    if args.force:
        # --force = generate ใหม่ทั้งหมด: เคลียร์ registry ที่โหลดจาก checkpoint เดิม
        ctx.event_times.clear()
        ctx.drama_descriptions.clear()
        ctx.drama_emitted.clear()
        ctx.pair_registry.clear()
        ctx.pair_emitted.clear()
        ctx.drama_pending.clear()
        ctx.drama_side_emitted.clear()
        ctx.drama_participant_cache.clear()
        ctx._desc_requested.clear()
        d = Path(ctx.config["output_dir"]).parent
        d.mkdir(parents=True, exist_ok=True)
        (d / "injected_events.jsonl").write_text("", encoding="utf-8")
        (d / "progress.jsonl").write_text("", encoding="utf-8")
        ctx.injected_events = []
        ctx.progress_log = []
    mode = "API (deepseek-v4-flash)" if ctx.client.using_api else "OFFLINE (template)"
    print(f"[phase2] mode={mode} | workers={args.workers} | target={args.target_rows} "
          f"| out={Path(ctx.config['output_dir'])}")

    if args.all:
        codes = sorted(ctx.identity_by_code)
    else:
        codes = [c.strip() for c in args.codes.split(",") if c.strip()]
    missing = [c for c in codes if c not in ctx.identity_by_code]
    if missing:
        print("[phase2] WARN: unknown codes (ข้าม):", missing)
        codes = [c for c in codes if c in ctx.identity_by_code]
    if not codes:
        print("[phase2] ไม่มีพนักงานให้ generate (ใช้ --codes หรือ --all)")
        return []

    def run(code: str) -> Dict[str, Any]:
        emp = ctx.identity_by_code[code]
        try:
            r = generate_employee(emp, ctx)
            r["name"] = emp["name"]
            r["department"] = emp.get("department", "")
            return r
        except Exception as e:  # บันทึก error ต่อคน ไม่ให้ทั้ง batch พัง
            return {"employeeCode": code, "error": f"{type(e).__name__}: {e}"}

    if args.workers > 1 and len(codes) > 1:
        with ThreadPoolExecutor(max_workers=args.workers) as ex:
            results = list(ex.map(run, codes))
    else:
        results = [run(c) for c in codes]

    # safety net: หลังวางแผนครบทุกคน — เขียนฝั่ง mirror ที่ยังขาด (กรณีวิ่ง parallel/ลำดับไม่ตรง)
    backfilled = ctx.backfill_drama_pending()
    if backfilled:
        print(f"[phase2] backfill: เขียน mirror ที่ขาดเพิ่ม {backfilled} rows")
    ctx.save_checkpoint()
    print("\n[phase2] === SUMMARY ===")
    for r in results:
        if r.get("error"):
            print(f"  {r['employeeCode']:8s} ERROR {r['error']}")
        elif r.get("skipped"):
            print(f"  {r['employeeCode']:8s} SKIP ({r.get('reason')})")
        else:
            print(f"  {r['employeeCode']:8s} +{r.get('rowsAdded',0):4d} rows "
                  f"(drama={r.get('dramaRows',0)} routine={r.get('routineRows',0)}) "
                  f"sheets={len(r.get('sheetsUpdated',[]))} final={r.get('finalRows',0)}")
    ok = sum(1 for r in results if not r.get("error") and not r.get("skipped"))
    print(f"[phase2] done: {ok}/{len(results)} generated | "
          f"total injected events in ctx: {len(ctx.injected_events)}")
    return results


if __name__ == "__main__":
    main()
